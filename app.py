# app.py (نسخه نهایی — کامل و یکپارچه - بهینه‌شده)
import os
import sys
import io
import re
import logging
import hashlib
import json
import sqlite3
import uuid
import unicodedata
from urllib.parse import urlparse, urljoin
from functools import wraps
from io import BytesIO
from logging.handlers import RotatingFileHandler
from datetime import datetime, timedelta, timezone

# بارگذاری متغیرهای محیطی از فایل .env (اگه وجود داشته باشه). این‌جوری لازم
# نیست SMTP_HOST/SMTP_USERNAME/SMTP_PASSWORD و... رو دستی در محیط ویندوز
# (setx / متغیر سیستمی) ست کنی؛ کافیه توی یک فایل .env کنار app.py بنویسی‌شون.
# اگه فایل .env وجود نداشته باشه، load_dotenv() بی‌خطر هیچ کاری نمی‌کنه و
# برنامه مثل قبل با متغیرهای محیطی واقعی سیستم کار می‌کنه.
from dotenv import load_dotenv
load_dotenv()

# مانیتورینگ خطا (اختیاری) — اگه پکیج نصب نباشه یا SENTRY_DSN ست نشده باشه،
# کاملاً بی‌اثر و بی‌خطره؛ فقط وقتی سایت آنلاین شد و SENTRY_DSN رو ست کردی فعال می‌شه.
try:
    import sentry_sdk
    from sentry_sdk.integrations.flask import FlaskIntegration
except ImportError:
    sentry_sdk = None

import jdatetime
import pytz
import pandas as pd
import plotly.express as px
import plotly.utils
from markupsafe import Markup
from hazm import Normalizer, word_tokenize
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from sqlalchemy import func, case, distinct
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate

from flask import (
    Flask, render_template, redirect, url_for, request, flash,
    send_from_directory, send_file, make_response, jsonify, abort
)

from flask_login import (
    LoginManager, login_user, logout_user, login_required,
    current_user, UserMixin
)

from flask_wtf import FlaskForm
from flask_wtf.csrf import CSRFProtect
from flask_wtf.file import FileField, FileAllowed
from wtforms import StringField, TextAreaField, SelectField, SubmitField, IntegerField
from wtforms.validators import DataRequired, Optional, Length, NumberRange

from payment_gateway import get_gateway, PaymentResult

from werkzeug.utils import secure_filename

from flask_bcrypt import Bcrypt  # ایمپورت Bcrypt
from flask_paginate import Pagination, get_page_parameter, get_per_page_parameter
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_caching import Cache
from pdf_encoding_fixer import process_pdf_on_upload
from mail_utils import (
    send_verification_email, send_password_reset_email,
    verify_token, EMAIL_VERIFY_SALT, PASSWORD_RESET_SALT,
)

# ---------- پیکربندی اولیه ----------
db = SQLAlchemy()
migrate = Migrate()
login_manager = LoginManager()
login_manager.login_view = 'login'
bcrypt = Bcrypt() # ایجاد شیء Bcrypt
# محافظت CSRF فقط برای مسیرهایی که از FlaskForm.validate_on_submit() استفاده
# می‌کنن به‌صورت خودکار اعمال می‌شه. مسیرهایی مثل /login، /register، /forgot،
# /toggle_bookmark، /add_bookmark و /save_reading_progress مستقیم از
# request.form/request.get_json می‌خونن و هیچ FlaskForm اعتبارسنجی نمی‌کنن،
# پس بدون این خط اصلاً محافظت CSRF نداشتن. CSRFProtect(app) این محافظت رو
# سراسری روی همه‌ی درخواست‌های POST/PUT/DELETE/PATCH اعمال می‌کنه.
csrf = CSRFProtect()
# محدودیت تعداد درخواست بر اساس IP، برای جلوگیری از حمله‌ی brute-force روی
# فرم‌های لاگین/ثبت‌نام/فراموشی رمز. مقدار پیش‌فرض سراسری سخت‌گیرانه نیست؛
# محدودیت واقعی روی هر روت حساس جداگانه با دکوریتور @limiter.limit ست می‌شه.
limiter = Limiter(key_func=get_remote_address, default_limits=[])

# حداکثر حجم مجاز به‌ازای نوع فایل (بایت). سقف کلی MAX_CONTENT_LENGTH در
# create_app هم به‌عنوان یک محدودیت سراسری سخت‌گیرانه‌تر اعمال می‌شه؛ این‌جا
# محدودیت دقیق‌تر بر اساس نوع فایل انجام می‌گیره.
FILE_SIZE_LIMITS = {
    '.pdf': 80 * 1024 * 1024,    # 80 مگابایت
    '.epub': 40 * 1024 * 1024,   # 40 مگابایت
    '.txt': 15 * 1024 * 1024,    # 15 مگابایت
    '.jpg': 5 * 1024 * 1024,     # 5 مگابایت
    '.jpeg': 5 * 1024 * 1024,
    '.png': 5 * 1024 * 1024,
}


def get_file_size(file_storage):
    """حجم فایل آپلودی (بایت) را بدون ذخیره‌ی کامل آن روی دیسک برمی‌گرداند."""
    stream = file_storage.stream
    stream.seek(0, os.SEEK_END)
    size = stream.tell()
    stream.seek(0)
    return size


def check_file_size_limit(file_storage, extension):
    """اگر حجم فایل بیشتر از سقف مجاز آن نوع فایل باشد، پیام خطا برمی‌گرداند؛
    در غیر این صورت None برمی‌گرداند."""
    limit = FILE_SIZE_LIMITS.get(extension.lower())
    if not limit:
        return None
    size = get_file_size(file_storage)
    if size > limit:
        limit_mb = limit // (1024 * 1024)
        return f'حجم فایل بیشتر از حد مجاز است (حداکثر {limit_mb} مگابایت برای فایل‌های {extension}).'
    return None


EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')


def normalize_persian_text(text):
    """نرمال‌سازی متن فارسی"""
    normalizer = Normalizer()
    return normalizer.normalize(text)

def extract_persian_text(text):
    """استخراج متن فارسی از متن ورودی"""
    # حذف کاراکترهای غیرفارسی و اعداد
    persian_text = re.sub(r'[^\u0600-\u06FF\s]', ' ', text)
    # حذف فاصله‌های اضافی
    return ' '.join(persian_text.split())

def process_pdf_content(content):
    """پردازش محتوای PDF برای نمایش بهتر متن فارسی"""
    # نرمال‌سازی متن
    normalized = normalize_persian_text(content)
    # استخراج متن فارسی
    persian_text = extract_persian_text(normalized)
    return persian_text
    
# ---------- فیلتر جینجا ----------
def timesince(dt, default="همین حالا"):
    if not dt:
        return default
    try:
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        now = datetime.now(timezone.utc)
        diff = now - dt

        periods = [
            (diff.days // 365, 'سال'),
            ((diff.days % 365) // 30, 'ماه'),
            ((diff.days % 30), 'روز'),
            (diff.seconds // 3600, 'ساعت'),
            ((diff.seconds % 3600) // 60, 'دقیقه'),
        ]

        for period, name in periods:
            if period:
                return f"{period} {name} پیش"
        return default
    except Exception:
        return default

# ---------- دکوراتور ادمین ----------
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not getattr(current_user, 'is_admin', False):
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

def is_safe_url(target):
    """بررسی安全 بودن URL برای جلوگیری از Open Redirect.
    فقط URL‌هایی را معتبر می‌داند که به همین هاست (هسته‌ی سایت) هستند و به
    protocol معتبر دارند و netloc برابر با هاست فعلی کاربر است. مثلاً جلوگیری از
    /login?next=https://evil.com
    """
    if not target:
        return False
    try:
        ref_url = urlparse(request.host_url)
        test_url = urlparse(urljoin(request.host_url, target))
        return (
            test_url.scheme in ('http', 'https') and
            ref_url.netloc == test_url.netloc
        )
    except Exception:
        return False

# ---------- فرم پروفایل ----------
CATEGORY_CHOICES = [
    ('general', 'عمومی'),
    ('novel', 'رمان'),
    ('science', 'علمی'),
    ('history', 'تاریخی'),
    ('religion', 'مذهبی'),
    ('historical_religious', 'تاریخی مذهبی'),
]
COVER_VALIDATORS = [FileAllowed(['jpg', 'jpeg', 'png'], 'فقط تصاویر مجاز هستند!')]
PRICE_VALIDATORS = [Optional(), NumberRange(min=0, message='قیمت نمی‌تواند منفی باشد')]

class BookUploadForm(FlaskForm):
    title = StringField('عنوان کتاب', validators=[DataRequired()])
    author = StringField('نویسنده', validators=[DataRequired()])
    description = TextAreaField('توضیحات', validators=[Optional()])
    category = SelectField('دسته‌بندی', choices=CATEGORY_CHOICES)
    price = IntegerField('قیمت (تومان) — برای رایگان بودن، صفر بگذارید', default=0,
                          validators=PRICE_VALIDATORS)
    file = FileField('فایل کتاب (PDF، TXT یا EPUB)', validators=[
    DataRequired(),
    FileAllowed(['pdf', 'txt', 'epub'], 'فقط فایل‌های PDF، TXT و EPUB مجاز هستند!')
    ])
    cover = FileField('تصویر جلد (اختیاری)', validators=COVER_VALIDATORS)
    submit = SubmitField('آپلود کتاب')

class EditBookForm(BookUploadForm):
    """فرم ویرایش کتاب — بدون فیلدهای description و file."""
    description = None
    file = None
    submit = SubmitField('ذخیره تغییرات')

class Message(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sender_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    recipient_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    content = db.Column(db.Text, nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    is_read = db.Column(db.Boolean, default=False)
    
    sender = db.relationship('User', foreign_keys=[sender_id], backref='sent_messages')
    recipient = db.relationship('User', foreign_keys=[recipient_id], backref='received_messages')

class ProfileForm(FlaskForm):
    full_name = StringField('نام کامل', validators=[DataRequired()])
    location = StringField('محل سکونت')
    job = StringField('شغل')
    education = StringField('تحصیلات')
    field_of_study = StringField('رشته تحصیلی')
    bio = TextAreaField('درباره من')
    interests = StringField('علاقه‌مندی‌ها (با کاما جدا کنید)')
    phone_number = StringField('شماره تلفن (11 رقمی - اختیاری)', validators=[Optional(), Length(min=11, max=11, message='شماره تلفن باید 11 رقمی باشد')])
    social_media = StringField('شبکه‌های اجتماعی (اختیاری)')
    email = StringField('ایمیل', validators=[Optional(), Length(max=255)])
    profile_image = FileField('تصویر پروفایل', validators=[FileAllowed(['jpg','jpeg','png'])])
    submit = SubmitField('ذخیره تغییرات')

class MessageForm(FlaskForm):
    content = TextAreaField('پیام', validators=[DataRequired()])
    submit = SubmitField('ارسال پیام')

# ---------- مدل‌ها ----------
class Bookmark(db.Model):
    __tablename__ = 'bookmarks'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    book_id = db.Column(db.Integer, db.ForeignKey('book.id'), nullable=False)
    page = db.Column(db.Integer, nullable=False)
    note = db.Column(db.String(200))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    user = db.relationship('User', back_populates='bookmarks')
    book = db.relationship('Book', back_populates='bookmarks')

class User(UserMixin, db.Model):
    __tablename__ = 'user'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    recovery_code = db.Column(db.String(200), nullable=False)
    # ایمیل کاربر: برای حساب‌های قدیمی که هنوز ایمیل ثبت نکرده‌اند nullable
    # است، ولی از ثبت‌نام‌های جدید به بعد اجباری است (چک در روت /register).
    # از این فیلد برای «بازیابی رمز عبور» و «تأیید ایمیل» استفاده می‌شود.
    email = db.Column(db.String(255), unique=True, nullable=True)
    email_verified = db.Column(db.Boolean, default=False, nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    last_login = db.Column(db.DateTime, nullable=True)
    last_read_book_id = db.Column(db.Integer, db.ForeignKey('book.id'), nullable=True)
    last_checked_books = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    profile = db.relationship('UserProfile', back_populates='user', uselist=False, cascade='all, delete-orphan')
    books = db.relationship('Book', back_populates='uploader', 
                          foreign_keys='Book.user_id',
                          lazy='dynamic')
    reading_progress = db.relationship('ReadingProgress', back_populates='user', lazy='dynamic', cascade='all, delete-orphan')
    bookmarks = db.relationship('Bookmark', back_populates='user', lazy='dynamic', cascade='all, delete-orphan')

class Book(db.Model):
    __tablename__ = 'book'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    title = db.Column(db.String(300), nullable=False, index=True)
    author = db.Column(db.String(200), nullable=True, index=True)
    description = db.Column(db.Text, nullable=True)  # توضیحات کتاب
    content = db.Column(db.Text, nullable=True)
    pdf_filename = db.Column(db.String(200), nullable=True)
    cover_filename = db.Column(db.String(200), nullable=True)
    category = db.Column(db.String(50), nullable=False, default='txt')
    price = db.Column(db.Integer, nullable=False, default=0)  # قیمت به تومان؛ 0 یعنی رایگان
    read_count = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    file_type = db.Column(db.String(10), default='pdf')  # نوع فایل: pdf, txt, epub
    
    # Relationships
    uploader = db.relationship('User', back_populates='books',
                             foreign_keys=[user_id])
    bookmarks = db.relationship('Bookmark', back_populates='book', lazy='dynamic', cascade='all, delete-orphan')
    progresses = db.relationship('ReadingProgress', back_populates='book', lazy='dynamic', cascade='all, delete-orphan')

class ReadingProgress(db.Model):
    __tablename__ = 'reading_progress'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    book_id = db.Column(db.Integer, db.ForeignKey('book.id'), nullable=False)
    position = db.Column(db.Integer, default=0)
    total_pages = db.Column(db.Integer, default=0)  # Total pages in the book
    completed = db.Column(db.Boolean, default=False)  # Whether the book has been completed
    last_opened = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    user = db.relationship('User', back_populates='reading_progress')
    book = db.relationship('Book', back_populates='progresses')

class Purchase(db.Model):
    """هر رکورد یعنی یک تلاش برای خرید یک کتاب (موفق یا ناموفق)."""
    __tablename__ = 'purchase'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    book_id = db.Column(db.Integer, db.ForeignKey('book.id'), nullable=False)
    amount = db.Column(db.Integer, nullable=False)  # تومان
    status = db.Column(db.String(20), nullable=False, default='pending')  # pending, paid, failed
    gateway = db.Column(db.String(20), nullable=False, default='sandbox')
    track_id = db.Column(db.String(120), nullable=True)   # شناسه‌ی تراکنش درگاه
    ref_number = db.Column(db.String(120), nullable=True)  # کد پیگیری بعد از پرداخت موفق
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    paid_at = db.Column(db.DateTime, nullable=True)

    user = db.relationship('User', backref=db.backref('purchases', lazy='dynamic'))
    book = db.relationship('Book', backref=db.backref('purchases', lazy='dynamic'))


class ActivityLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    action = db.Column(db.String(100))
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

class UserProfile(db.Model):
    __tablename__ = 'user_profile'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), unique=True)
    full_name = db.Column(db.String(100))
    location = db.Column(db.String(100))
    job = db.Column(db.String(100))
    education = db.Column(db.String(100))
    field_of_study = db.Column(db.String(100))
    bio = db.Column(db.Text)
    phone_number = db.Column(db.String(20), nullable=True)
    social_media = db.Column(db.String(100))
    interests = db.Column(db.String(500))
    # این ستون قبلاً در مدل وجود نداشت با این‌که کد در چند جا (روت /profile و
    # قالب‌ها) روی profile.profile_image می‌نوشت/می‌خواند — یعنی مقدار تصویر
    # پروفایل هیچ‌وقت واقعاً ذخیره نمی‌شد. اضافه شدنش این باگ رو رفع می‌کنه.
    profile_image = db.Column(db.String(255), nullable=True)
    
    # Relationships
    user = db.relationship('User', back_populates='profile')
    
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# ---------- توابع کمکی ----------
def find_books_with_persian_titles():
    """پیدا کردن کتاب‌هایی که عنوان فارسی دارند"""
    all_books = Book.query.all()
    persian_books = []
    
    for book in all_books:
        if book.title:
            try:
                # تبدیل به یونیکد و بررسی وجود حروف فارسی
                if any('\u0600' <= char <= '\u06FF' for char in book.title):
                    print(f"عنوان فارسی یافت شد: '{book.title}'")
                    persian_books.append(book)
            except Exception as e:
                print(f"خطا در پردازش عنوان: {book.title} - {str(e)}")
                continue
    
    print(f"تعداد کتاب‌های با عنوان فارسی: {len(persian_books)}")
    return persian_books

def ensure_default_admin():
    try:
        if not User.query.filter_by(is_admin=True).first():
            # رمز عبور و کد بازیابی ادمین باید از متغیرهای محیطی خونده بشن.
            # اگه ست نشده باشن، یک مقدار تصادفی امن تولید و در لاگ چاپ می‌شه
            # تا هیچ‌وقت یک پسورد ثابت و قابل‌حدس در کد نباشه.
            admin_username = os.environ.get('ADMIN_USERNAME', 'admin')
            admin_password_plain = os.environ.get('ADMIN_PASSWORD')
            admin_recovery_plain = os.environ.get('ADMIN_RECOVERY_CODE')

            generated_password = False
            if not admin_password_plain:
                admin_password_plain = uuid.uuid4().hex
                generated_password = True
            if not admin_recovery_plain:
                admin_recovery_plain = uuid.uuid4().hex
                generated_password = True

            # هش کردن رمز عبور با Bcrypt
            admin_password = bcrypt.generate_password_hash(admin_password_plain).decode('utf-8')
            admin_recovery = bcrypt.generate_password_hash(admin_recovery_plain).decode('utf-8')
            admin = User(
                username=admin_username,
                password_hash=admin_password,
                recovery_code=admin_recovery,
                is_admin=True,
                last_read_book_id=None
            )
            db.session.add(admin)
            db.session.commit()

            if generated_password:
                # این پیام فقط یک‌بار موقع ساخت ادمین پیش‌فرض چاپ می‌شه.
                # حتماً این مقادیر رو یادداشت کن و پسورد رو بلافاصله بعد از اولین ورود تغییر بده.
                print(
                    f"[ensure_default_admin] کاربر ادمین پیش‌فرض ساخته شد.\n"
                    f"  username: {admin_username}\n"
                    f"  password: {admin_password_plain}\n"
                    f"  recovery_code: {admin_recovery_plain}\n"
                    f"برای جلوگیری از این پیام، متغیرهای محیطی ADMIN_USERNAME، "
                    f"ADMIN_PASSWORD و ADMIN_RECOVERY_CODE را تنظیم کنید."
                )
    except Exception as e:
        if 'no such column' in str(e) and 'last_read_book_id' in str(e):
            pass
        else:
            raise

def _file_hash_text(text: str) -> str:
    return hashlib.md5((text or '').encode('utf-8')).hexdigest()

def scan_books_folder(BASE_DIR):
    stats = {'added_pdf': 0, 'removed_pdf': 0, 'added_txt': 0, 'removed_txt': 0, 'updated_txt': 0, 'added_cover': 0}
    pdf_path = os.path.join(BASE_DIR, 'all_books', 'pdf')
    db_pdfs = {b.pdf_filename: b for b in Book.query.filter_by(category='pdf').all() if b.pdf_filename}
    fs_pdfs = set()
    
    if os.path.exists(pdf_path):
        for filename in os.listdir(pdf_path):
            if not filename.lower().endswith('.pdf'):
                continue
            fs_pdfs.add(filename)
            if filename not in db_pdfs:
                title = os.path.splitext(filename)[0]
                book = Book(title=title, author='ناشناس', pdf_filename=filename, category='pdf', file_type='pdf')
                cover_path = os.path.join(BASE_DIR, 'all_books', 'covers', f"{title}.jpg")
                if os.path.exists(cover_path):
                    book.cover_filename = f"{title}.jpg"
                    stats['added_cover'] += 1
                db.session.add(book)
                stats['added_pdf'] += 1
    
    for filename, book in list(db_pdfs.items()):
        if filename not in fs_pdfs:
            ReadingProgress.query.filter_by(book_id=book.id).delete()
            db.session.delete(book)
            stats['removed_pdf'] += 1

    txt_path = os.path.join(BASE_DIR, 'all_books', 'txt')
    db_txts = {b.title: b for b in Book.query.filter_by(category='txt').all()}
    fs_txts = set()
    
    if os.path.exists(txt_path):
        for filename in os.listdir(txt_path):
            if not filename.lower().endswith('.txt'):
                continue
            fs_txts.add(filename)
            title = os.path.splitext(filename)[0]
            file_full = os.path.join(txt_path, filename)
            
            try:
                with open(file_full, 'r', encoding='utf-8') as f:
                    content = f.read()
            except Exception:
                with open(file_full, 'r', encoding='latin-1') as f:
                    content = f.read()
            
            if title not in db_txts:
                book = Book(title=title, author='ناشناس', content=content, category='txt', file_type='txt')
                cover_path = os.path.join(BASE_DIR, 'all_books', 'covers', f"{title}.jpg")
                if os.path.exists(cover_path):
                    book.cover_filename = f"{title}.jpg"
                    stats['added_cover'] += 1
                db.session.add(book)
                stats['added_txt'] += 1
            else:
                book = db_txts[title]
                if _file_hash_text(book.content or '') != _file_hash_text(content or ''):
                    book.content = content
                    stats['updated_txt'] += 1
    
    for title, book in list(db_txts.items()):
        if f"{title}.txt" not in fs_txts:
            ReadingProgress.query.filter_by(book_id=book.id).delete()
            db.session.delete(book)
            stats['removed_txt'] += 1
    
    db.session.commit()

    return stats

def upgrade_database(db_path):
    # safe upgrade: work only if db exists and book table exists
    if not os.path.exists(db_path):
        return
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='book';")
        if not cursor.fetchone():
            conn.commit()
            return

        cursor.execute("PRAGMA table_info(book);")
        cols = cursor.fetchall()
        columns = {c[1]: c for c in cols}

        # if user_id exists and NOT NULL -> rebuild
        if "user_id" in columns:
            cid, name, col_type, notnull, default, pk = columns["user_id"]
            if notnull == 1:
                cursor.execute("ALTER TABLE book RENAME TO book_old;")
                cursor.execute("""
                CREATE TABLE book (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    title TEXT NOT NULL,
                    author TEXT,
                    content TEXT,
                    pdf_filename TEXT,
                    cover_filename TEXT,
                    category TEXT NOT NULL,
                    read_count INTEGER DEFAULT 0,
                    FOREIGN KEY(user_id) REFERENCES user(id)
                );
                """)
                cursor.execute("""
                INSERT INTO book (id, user_id, title, author, content, pdf_filename, cover_filename, category, read_count)
                SELECT id, NULL, title, author, content, pdf_filename, cover_filename, category, read_count
                FROM book_old;
                """)
                cursor.execute("DROP TABLE book_old;")

        # add cover_filename if missing
        cursor.execute("PRAGMA table_info(book);")
        cols2 = [c[1] for c in cursor.fetchall()]
        if "cover_filename" not in cols2:
            try:
                cursor.execute("ALTER TABLE book ADD COLUMN cover_filename TEXT;")
            except Exception:
                pass

        # add price if missing (برای فعال‌سازی فروش کتاب روی نسخه‌های قدیمی‌تر پایگاه‌داده)
        cursor.execute("PRAGMA table_info(book);")
        cols3 = [c[1] for c in cursor.fetchall()]
        if "price" not in cols3:
            try:
                cursor.execute("ALTER TABLE book ADD COLUMN price INTEGER NOT NULL DEFAULT 0;")
            except Exception:
                pass

        # ---- جدول user: اضافه کردن ستون‌های تأیید ایمیل روی دیتابیس‌های قدیمی ----
        # این ستون‌ها با مایگریشن a1b2c3d4e5f6 اضافه شدن، ولی چون این تابع
        # (upgrade_database) مسیر سبکی است که بدون نیاز به alembic اجرا می‌شه،
        # باید این‌جا هم دستی چک/اضافه بشن؛ وگرنه روی دیتابیس‌های قدیمی که
        # هنوز این ستون‌ها رو ندارن، همون اولین SELECT روی جدول user (مثلاً
        # موقع لاگین) با خطای "no such column: user.email" شکست می‌خوره.
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='user';")
        if cursor.fetchone():
            cursor.execute("PRAGMA table_info(user);")
            user_cols = [c[1] for c in cursor.fetchall()]
            if "email" not in user_cols:
                try:
                    cursor.execute("ALTER TABLE user ADD COLUMN email VARCHAR(255);")
                except Exception:
                    pass
            if "email_verified" not in user_cols:
                try:
                    cursor.execute("ALTER TABLE user ADD COLUMN email_verified BOOLEAN NOT NULL DEFAULT 0;")
                except Exception:
                    pass

        # ---- جدول user_profile: اضافه کردن ستون تصویر پروفایل روی دیتابیس‌های قدیمی ----
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='user_profile';")
        if cursor.fetchone():
            cursor.execute("PRAGMA table_info(user_profile);")
            profile_cols = [c[1] for c in cursor.fetchall()]
            if "profile_image" not in profile_cols:
                try:
                    cursor.execute("ALTER TABLE user_profile ADD COLUMN profile_image VARCHAR(255);")
                except Exception:
                    pass
            if "phone_number" not in profile_cols:
                try:
                    cursor.execute("ALTER TABLE user_profile ADD COLUMN phone_number VARCHAR(20);")
                except Exception:
                    pass

        conn.commit()
    finally:
        conn.close()

# ---------- ساخت اپ ----------
def number_format(value, decimal_places=0, decimal_sep='.', thousands_sep=','):
    """Format a number with thousands separator and decimal places"""
    try:
        value = float(value)
        # Format the number with thousands separator and decimal places
        parts = f"{value:,.{decimal_places}f}".split(".")
        integer_part = parts[0].replace(",", thousands_sep)
        if len(parts) > 1:
            return f"{integer_part}{decimal_sep}{parts[1]}"
        return integer_part
    except (ValueError, TypeError):
        return str(value)

TEHRAN_TZ = pytz.timezone('Asia/Tehran')

def _to_tehran_time(dt):
    """دیتابیس زمان‌ها رو به‌صورت UTC ساده (بدون tzinfo) ذخیره می‌کنه
    (datetime.utcnow()). قبل از نمایش باید به وقت محلی ایران تبدیل بشه،
    وگرنه گزارش‌ها ساعت UTC رو نشون می‌دن نه ساعت واقعی خرید/رویداد.
    در انتها tzinfo رو حذف می‌کنیم چون jdatetime با datetime آگاه از
    منطقه‌زمانی (tz-aware) درست کار نمی‌کنه و بی‌صدا خطا می‌ده."""
    if dt.tzinfo is None:
        dt = pytz.utc.localize(dt)
    return dt.astimezone(TEHRAN_TZ).replace(tzinfo=None)

def format_jalali(dt, default=""):
    if not dt:
        return default
    try:
        dt = _to_tehran_time(dt)
        jalali_date = jdatetime.datetime.fromgregorian(datetime=dt)
        return jalali_date.strftime('%Y/%m/%d %H:%M')
    except Exception as e:
        try:
            logging.getLogger(__name__).error(
                f"format_jalali error for dt={dt!r}: {type(e).__name__}: {e}"
            )
        except Exception:
            pass
        return default

def create_app():
    app = Flask(__name__, static_folder='static', static_url_path='/static')
    
    # Add Jalali date filter to Jinja2
    app.jinja_env.filters['jalali'] = format_jalali
    if not app.debug:
        if not os.path.exists('logs'):
            os.mkdir('logs')
        file_handler = RotatingFileHandler('logs/digital_library.log', maxBytes=10240, backupCount=10)
        file_handler.setFormatter(logging.Formatter(
            '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
        ))
        file_handler.setLevel(logging.INFO)
        app.logger.addHandler(file_handler)
        app.logger.setLevel(logging.INFO)
        app.logger.info('Digital Library startup')
    
    # APP_ENV مشخص می‌کنه محیط اجرا production هست یا نه. هر مقداری غیر از
    # 'production' (پیش‌فرض 'development') به‌عنوان محیط توسعه در نظر گرفته می‌شه.
    app_env = os.environ.get('APP_ENV', os.environ.get('FLASK_ENV', 'development')).lower()
    is_production = app_env == 'production'

    # مانیتورینگ خطا با Sentry — کاملاً اختیاری. فقط وقتی هم پکیج نصب باشه هم
    # SENTRY_DSN ست شده باشه فعال می‌شه؛ در غیر این صورت هیچ اتفاقی نمی‌افته.
    sentry_dsn = os.environ.get('SENTRY_DSN')
    if sentry_dsn and sentry_sdk is not None:
        sentry_sdk.init(
            dsn=sentry_dsn,
            integrations=[FlaskIntegration()],
            environment=app_env,
            traces_sample_rate=0.1,
        )

    secret_key = os.environ.get('SECRET_KEY')
    if not secret_key:
        if is_production:
            # در production هرگز نباید با یک SECRET_KEY تصادفی بالا بیاد: با هر
            # ری‌استارت/دیپلوی همه‌ی session‌ها و CSRF tokenها باطل می‌شن. به‌جای
            # هشدار و ادامه‌ی کار، اجرای برنامه رو کاملاً متوقف می‌کنیم.
            raise RuntimeError(
                "SECRET_KEY تنظیم نشده و APP_ENV=production است. برای اجرای "
                "امن در production باید یک مقدار ثابت و تصادفی (مثلاً خروجی "
                "`python -c \"import secrets; print(secrets.token_hex(32))\"`) "
                "را در متغیر محیطی SECRET_KEY تنظیم کنید."
            )
        # هرگز از یک مقدار ثابت به‌عنوان SECRET_KEY پیش‌فرض استفاده نکن؛ این مقدار
        # امنیت session و CSRF token‌ها رو تضمین می‌کنه. اگه از env ست نشده باشه،
        # یک مقدار تصادفی امن در هر اجرا تولید می‌شه (توجه: با ری‌استارت شدن،
        # تمام session‌های قبلی نامعتبر می‌شن؛ برای پروداکشن حتماً SECRET_KEY رو
        # به‌صورت ثابت در متغیر محیطی تنظیم کن).
        secret_key = uuid.uuid4().hex + uuid.uuid4().hex
        print(
            "[create_app] هشدار: متغیر محیطی SECRET_KEY تنظیم نشده. "
            "یک کلید تصادفی موقت تولید شد. برای محیط پروداکشن حتماً "
            "SECRET_KEY را در متغیرهای محیطی تنظیم کنید (و APP_ENV=production "
            "را هم ست کنید تا این هشدار جدی گرفته بشه)."
        )
    app.config['SECRET_KEY'] = secret_key
    app.config['APP_ENV'] = app_env
    app.config['IS_PRODUCTION'] = is_production
    BASE_DIR = os.path.abspath(os.path.dirname(__file__))
    app.config['BASE_DIR'] = BASE_DIR
    app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
        'DATABASE_URL', 'sqlite:///' + os.path.join(BASE_DIR, 'data', 'library.db')
    )
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

    # ---------- تنظیمات امنیتی کوکی سشن ----------
    # SESSION_COOKIE_SECURE یعنی کوکی فقط روی HTTPS ارسال می‌شه؛ در محیط توسعه
    # (بدون HTTPS محلی) به‌صورت پیش‌فرض خاموشه مگر این‌که صراحتاً روشن بشه.
    session_cookie_secure = os.environ.get('SESSION_COOKIE_SECURE')
    if session_cookie_secure is None:
        session_cookie_secure = is_production
    else:
        session_cookie_secure = session_cookie_secure.lower() in ('1', 'true', 'yes')
    app.config['SESSION_COOKIE_SECURE'] = session_cookie_secure
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
    app.config['REMEMBER_COOKIE_SECURE'] = session_cookie_secure
    app.config['REMEMBER_COOKIE_HTTPONLY'] = True
    app.config['REMEMBER_COOKIE_SAMESITE'] = 'Lax'

    # ---------- محدودیت حجم آپلود ----------
    # این یه سقف کلی و سخت‌گیرانه‌ست (برای جلوگیری از DoS با آپلود فایل‌های
    # حجیم)؛ محدودیت دقیق‌تر بر اساس نوع فایل (PDF/TXT/EPUB/تصویر) در روت
    # /upload با تابع enforce_file_size_limit اعمال می‌شه.
    app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 مگابایت سقف کلی

    # ---------- تنظیمات ایمیل (SMTP) ----------
    # برای فعال کردن ارسال واقعی ایمیل (تأیید ثبت‌نام / بازیابی رمز)، متغیرهای
    # محیطی زیر را تنظیم کنید. تا زمانی که تنظیم نشده باشن، متن ایمیل‌ها فقط
    # در لاگ نوشته می‌شه (مناسب توسعه/تست) و کاربر پیام مناسب می‌بینه.
    app.config['SMTP_HOST'] = os.environ.get('SMTP_HOST')
    app.config['SMTP_PORT'] = int(os.environ.get('SMTP_PORT', '587'))
    app.config['SMTP_USERNAME'] = os.environ.get('SMTP_USERNAME')
    app.config['SMTP_PASSWORD'] = os.environ.get('SMTP_PASSWORD')
    app.config['SMTP_USE_TLS'] = os.environ.get('SMTP_USE_TLS', '1').lower() in ('1', 'true', 'yes')
    app.config['MAIL_SENDER'] = os.environ.get('MAIL_SENDER', app.config['SMTP_USERNAME'] or 'no-reply@digital-library.local')
    app.config['SITE_BASE_URL'] = os.environ.get('SITE_BASE_URL', '').rstrip('/')

    # ---------- تنظیمات درگاه پرداخت ----------
    # PAYMENT_MODE = 'sandbox' یعنی بدون درگاه واقعی (برای تست کامل مسیر خرید).
    # وقتی آماده‌ی اتصال واقعی به زیبال بودید، فقط این مقدار رو (یا متغیر محیطی
    # PAYMENT_MODE) به 'zibal' تغییر بدید و ZIBAL_MERCHANT رو با مرچنت واقعی خودتون
    # پر کنید — هیچ تغییر دیگه‌ای در کد لازم نیست.
    app.config['PAYMENT_MODE'] = os.environ.get('PAYMENT_MODE', 'sandbox')
    app.config['ZIBAL_MERCHANT'] = os.environ.get('ZIBAL_MERCHANT', 'zibal')  # 'zibal' = مرچنت آزمایشی رسمی
    
    # Define base upload directories
    app.config['UPLOAD_BASE'] = os.path.join(BASE_DIR, 'uploads')
    app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'uploads')
    app.config['COVER_FOLDER'] = os.path.join(BASE_DIR, 'uploads', 'covers')
    
    # Ensure all necessary directories exist
    required_dirs = [
        os.path.join(BASE_DIR, 'static', 'profile_images'),
        os.path.join(BASE_DIR, 'data'),
        os.path.join(BASE_DIR, 'static', 'uploads', 'charts'),
        app.config['UPLOAD_BASE'],
        app.config['UPLOAD_FOLDER'],
        app.config['COVER_FOLDER']
    ]
    
    for directory in required_dirs:
        os.makedirs(directory, exist_ok=True)

    db.init_app(app)
    migrate.init_app(app, db)
    login_manager.init_app(app)
    login_manager.login_view = 'login'
    login_manager.login_message = 'لطفاً برای مشاهده این صفحه وارد شوید'
    login_manager.login_message_category = 'warning'
    bcrypt.init_app(app)
    csrf.init_app(app)
    limiter.init_app(app)

    # Define the to_jalali function
    def to_jalali(dt, format_str='%Y/%m/%d %H:%M:%S'):
        """Convert datetime to Jalali with 24-hour time format (e.g., '1402/09/22 14:30:45')"""
        if not dt:
            if hasattr(app, 'logger'):
                app.logger.debug("No date provided to to_jalali")
            return "-"
        try:
            local_dt = _to_tehran_time(dt)
            jd = jdatetime.datetime.fromgregorian(datetime=local_dt)
            result = jd.strftime(format_str)
            if hasattr(app, 'logger'):
                app.logger.debug(f"Converted {dt} to Jalali: {result}")
            return result
        except Exception as e:
            error_msg = f"Error converting date {dt} to Jalali: {str(e)}"
            if hasattr(app, 'logger'):
                app.logger.error(error_msg)
            # Fallback to Gregorian if conversion fails
            return dt.strftime(format_str)

    # Register the filter
    # Register as both a global function and a filter
    app.jinja_env.globals['to_jalali'] = to_jalali
    app.jinja_env.filters['to_jalali'] = to_jalali

    with app.app_context():
        db_path = os.path.join(BASE_DIR, 'data', 'library.db')
        db.create_all()
        upgrade_database(db_path)
        db.create_all()
        ensure_default_admin()
        # Removed automatic book scanning on startup to prevent duplicate imports
        # scan_books_folder(BASE_DIR)  # Uncomment and use this manually when needed

    def allowed_file(filename, allowed_extensions=None):
        if allowed_extensions is None:
            allowed_extensions = {'pdf', 'txt'}
        return '.' in filename and \
            filename.rsplit('.', 1)[1].lower() in allowed_extensions

    def _strip_tz(dt):
        """تبدیل datetime به naive UTC بدون منطقه‌زمانی، برای سازگاری با مقادیر ذخیره‌شده در دیتابیس."""
        if dt is None:
            return datetime.min
        if dt.tzinfo is not None:
            dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
        return dt

    TYPE_TO_DIR = {'pdf': 'pdf', 'txt': 'txt', 'epub': 'epub'}
    TYPE_TO_MIME = {
        'pdf': 'application/pdf',
        'txt': 'text/plain; charset=utf-8',
        'epub': 'application/epub+zip',
    }

    def get_book_file_path(book):
        """مسیر فیزیکی فایل کتاب روی دیسک را بر اساس file_type برمی‌گرداند.
        نام فایل در فیلد `pdf_filename` نگه‌داری می‌شود (برای همه نوع فایل)."""
        if not book or not book.pdf_filename or not book.file_type:
            return None
        ft = book.file_type.lower()
        subdir = TYPE_TO_DIR.get(ft)
        if not subdir:
            return None
        return os.path.join(app.config['UPLOAD_FOLDER'], 'books', subdir, book.pdf_filename)

    def get_mime_for_book_type(file_type):
        if not file_type:
            return 'application/octet-stream'
        return TYPE_TO_MIME.get(file_type.lower(), 'application/octet-stream')

    def get_new_books_count(user):
        if not user.is_authenticated:
            return 0
        last_checked = _strip_tz(user.last_checked_books)
        return Book.query.filter(Book.created_at > last_checked).count()

    def get_unread_messages_count(user_id):
        if not user_id:
            return 0
        return Message.query.filter_by(recipient_id=user_id, is_read=False).count()

    def user_has_access(book):
        """آیا کاربر جاری اجازه دارد محتوای این کتاب را بخواند/دانلود کند؟
        کتاب‌های رایگان برای همه آزادند؛ کتاب‌های قیمت‌دار فقط برای ادمین،
        آپلودکننده، و کسی که خریدِ «paid» ثبت‌شده دارد."""
        if not book.price or book.price <= 0:
            return True
        if not current_user.is_authenticated:
            return False
        if current_user.is_admin or book.user_id == current_user.id:
            return True
        return Purchase.query.filter_by(
            user_id=current_user.id, book_id=book.id, status='paid'
        ).first() is not None

    @app.template_filter('comma_separated')
    def comma_separated(value):
        """Format number with comma as thousand separator"""
        if value is None:
            return "0"
        return f"{int(value):,}".replace(',', '،')  # Persian comma

    @app.template_filter('timesince')
    def timesince_filter(dt, default="همین حالا"):
        """Return a friendly time difference string"""
        if not dt:
            return default
        try:
            # اطمینان از مقایسه‌ی دو مقدار naive UTC با هم، حتی وقتی dt دارای tzinfo باشه
            dt_naive = _strip_tz(dt)
            diff = datetime.utcnow() - dt_naive
        except (TypeError, ValueError):
            return default
            
        if diff.days > 0:
            if diff.days == 1:
                return 'دیروز'
            elif diff.days < 7:
                return f'{diff.days} روز پیش'
            elif diff.days < 30:
                weeks = diff.days // 7
                return f'{weeks} هفته پیش'
            elif diff.days < 365:
                months = diff.days // 30
                return f'{months} ماه پیش'
            else:
                years = diff.days // 365
                return f'{years} سال پیش'
        else:
            if diff.seconds < 60:
                return 'چند لحظه پیش'
            elif diff.seconds < 3600:
                minutes = diff.seconds // 60
                return f'{minutes} دقیقه پیش'
            else:
                hours = diff.seconds // 3600
                return f'{hours} ساعت پیش'

    @app.template_filter('nl2br')
    def nl2br_filter(value):
        """Convert newlines to <br> tags for HTML display"""
        if value is None:
            return ''
        # ابتدا محتوای ورودی (که می‌تونه از کاربر باشه، مثل bio یا پیام‌ها) باید
        # escape بشه تا اگه شامل تگ‌های HTML/اسکریپت بود، به‌عنوان کد اجرا نشه.
        # بعد فقط <br> واقعی رو اضافه می‌کنیم و نتیجه رو Markup می‌کنیم تا Jinja
        # این <br>های خودمون رو دوباره escape نکنه، ولی محتوای اصلی کاربر
        # همچنان escape شده باقی می‌مونه.
        escaped = Markup.escape(str(value))
        return Markup(escaped.replace('\n', Markup('<br>')))

    @app.context_processor
    def inject_functions():
        def _get_new_books_count(_=None):
            return get_new_books_count(current_user._get_current_object())
            
        def _get_unread_messages_count(_=None):
            if not current_user.is_authenticated:
                return 0
            return get_unread_messages_count(current_user._get_current_object().id)

        return dict(
            get_new_books_count=_get_new_books_count,
            get_unread_messages_count=_get_unread_messages_count,
            nl2br=nl2br_filter,
            timesince=timesince_filter,
            user_has_access=user_has_access
        )

    # ---------- ریدایرکت برای مدیریت هدایت به داشبورد مناسب ----------
    @app.route('/upload', methods=['GET', 'POST'])
    @login_required
    def upload_book():
        form = BookUploadForm()
        
        if form.validate_on_submit():
            try:
                file = form.file.data
                title = form.title.data.strip()
                author = form.author.data.strip()
                description = form.description.data.strip()
                category = form.category.data
                
                # Generate a unique filename to prevent collisions
                filename = unicodedata.normalize('NFC', file.filename)  # نرمال‌سازی یونیکد نام فایل
                safe_filename = secure_filename(filename)  # حذف کاراکترهای نامناسب
                if not safe_filename:
                    safe_filename = f"book_{uuid.uuid4().hex}"
                unique_filename = f"{uuid.uuid4().hex}_{safe_filename}"
                
                # Check file extension
                file_parts = os.path.splitext(filename)
                if len(file_parts) > 1:
                    file_extension = file_parts[1].lower()
                else:
                    flash('فایل باید پسوند داشته باشد (مانند .pdf، .txt یا .epub)', 'error')
                    return render_template('upload_book.html', form=form)

                # بررسی حجم فایل بر اساس نوع آن (جلوگیری از آپلود فایل‌های
                # حجیم غیرمنطقی که می‌تونه دیسک/حافظه‌ی سرور رو پر کنه)
                size_error = check_file_size_limit(file, file_extension)
                if size_error:
                    flash(size_error, 'error')
                    return render_template('upload_book.html', form=form)

                # Determine the target directory based on file type
                if file_extension == '.pdf':
                    target_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'books', 'pdf')
                elif file_extension == '.txt':
                    target_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'books', 'txt')
                elif file_extension == '.epub':
                    target_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'books', 'epub')
                else:
                    flash('فرمت فایل پشتیبانی نمی‌شود', 'error')
                    return render_template('upload_book.html', form=form)
                
                # Create the target directory if it doesn't exist
                os.makedirs(target_dir, exist_ok=True)
                
                # Save the uploaded file to the appropriate directory
                file_path = os.path.join(target_dir, unique_filename)
                file.save(file_path)
                
                # Fix PDF encoding issues if this is a PDF file
                if file_extension == '.pdf':
                    try:
                        process_pdf_on_upload(file_path)
                    except Exception as e:
                        app.logger.error(f"Error processing PDF encoding: {e}")
                        # Continue with upload even if encoding fix fails

                # Handle cover image
                cover_filename = None
                if form.cover.data:
                    cover_file = form.cover.data
                    original_cover_name = secure_filename(cover_file.filename)
                    if '.' in original_cover_name:
                        cover_extension = original_cover_name.rsplit('.', 1)[1].lower()
                        if cover_extension not in ['jpg', 'jpeg', 'png']:
                            flash('فقط فایل‌های تصویری با پسوند jpg، jpeg یا png مجاز هستند', 'error')
                            return render_template('upload_book.html', form=form)
                    else:
                        flash('فایل تصویری باید پسوند معتبر داشته باشد', 'error')
                        return render_template('upload_book.html', form=form)

                    cover_size_error = check_file_size_limit(cover_file, f'.{cover_extension}')
                    if cover_size_error:
                        flash(cover_size_error, 'error')
                        return render_template('upload_book.html', form=form)

                    cover_filename = f"cover_{uuid.uuid4().hex}.{cover_extension}"
                    cover_path = os.path.join(app.config['UPLOAD_FOLDER'], 'covers', cover_filename)
                    # Create the covers directory if it doesn't exist
                    os.makedirs(os.path.dirname(cover_path), exist_ok=True)
                    cover_file.save(cover_path)

                # Handle different file types
                content = None
                if file_extension == '.txt':
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            content = f.read()
                    except Exception as e:
                        flash(f'خطا در خواندن فایل متنی: {str(e)}', 'error')
                        return render_template('upload_book.html', form=form)
                elif file_extension == '.epub':
                    try:
                        from epub_utils import EPUBReader
                        reader = EPUBReader(file_path)
                        metadata = reader.read_epub()
                        if metadata and 'content' in metadata:
                            content = metadata['content']
                            # Update title and author from EPUB metadata if available
                            if 'title' in metadata and not form.title.data:
                                form.title.data = metadata['title']
                            if 'author' in metadata and not form.author.data:
                                form.author.data = metadata['author']
                    except Exception as e:
                        flash(f'خطا در پردازش فایل EPUB: {str(e)}', 'error')
                        return render_template('upload_book.html', form=form)

                # Create new book record
                book = Book(
                    title=form.title.data,
                    author=form.author.data,
                    description=form.description.data or None,
                    category=form.category.data or 'general',
                    price=form.price.data or 0,
                    user_id=current_user.id,
                    pdf_filename=unique_filename,
                    cover_filename=cover_filename,
                    content=content,
                    file_type=file_extension[1:]  # Remove the dot from extension
                )
                db.session.add(book)
                db.session.commit()
                
                flash('کتاب با موفقیت آپلود شد', 'success')
                return redirect(url_for('books'))
                
            except Exception as e:
                db.session.rollback()
                # Clean up uploaded files if there was an error
                if 'file_path' in locals() and os.path.exists(file_path):
                    os.remove(file_path)
                if 'cover_filename' in locals() and cover_filename and os.path.exists(os.path.join(app.config['COVER_FOLDER'], cover_filename)):
                    os.remove(os.path.join(app.config['COVER_FOLDER'], cover_filename))
                
                app.logger.error(f"Error uploading book: {str(e)}")
                flash(f'خطا در آپلود کتاب: {str(e)}', 'error')
        
        # For GET request or failed validation, show the upload form
        return render_template('upload_book.html', form=form)  # Pass form to template


    @app.route('/books/persian-titles')
    @login_required
    def show_persian_title_books():
        """نمایش کتاب‌هایی که عنوان فارسی دارند"""
        all_books = Book.query.all()
        print(f"Total books in database: {len(all_books)}")
        
        persian_books = find_books_with_persian_titles()
        print(f"Found {len(persian_books)} persian books")
        
        # Print titles of all books for debugging
        for i, book in enumerate(all_books, 1):
            print(f"Book {i}: '{book.title}' - Type: {type(book.title)}")
        
        return render_template('all_books.html', 
                            books=persian_books, 
                            title='کتاب‌های با عنوان فارسی',
                            show_upload_button=current_user.is_authenticated)

    @app.route('/dashboard')
    @login_required
    def dashboard_redirect():
        if current_user.is_admin:
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('user_dashboard'))

    # ---------- index و داشبورد ----------
    @app.route('/')
    def index():
        if current_user.is_authenticated:
            return redirect(url_for('dashboard_redirect'))
        return redirect(url_for('login'))

    def get_pdf_page_count(book):
        """تعداد صفحات فایل را برمی‌گرداند بر اساس نوع فایل"""
        try:
            file_path = get_book_file_path(book)
            if not file_path:
                return 100
            ft = book.file_type.lower()
            if ft == 'pdf':
                import PyPDF2
                with open(file_path, 'rb') as f:
                    pdf_reader = PyPDF2.PdfReader(f)
                    return len(pdf_reader.pages)
            elif ft == 'txt':
                with open(file_path, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                    return max(1, len(lines) // 50)
            elif ft == 'epub':
                from epub_utils import EPUBReader
                reader = EPUBReader(file_path)
                metadata = reader.read_epub()
                if metadata and 'page_count' in metadata:
                    return metadata['page_count']
                return 100
        except Exception as e:
            app.logger.error(f"Error getting page count for {getattr(book, 'file_type', None)} {getattr(book, 'id', None)}: {str(e)}")
            return 100

    def estimate_total_pages(book):
        """تخمین تعداد صفحات کتاب بر اساس نوع محتوا"""
        if book.pdf_filename:
            # استفاده از تابع به‌روزرسانی شده get_pdf_page_count
            return get_pdf_page_count(book)
            
        # اگر محتوای متنی وجود دارد، بر اساس تعداد کاراکترها تخمین بزن
        if book.content:
            # میانگین تعداد کاراکتر در هر صفحه کتاب (تقریبی)
            CHARS_PER_PAGE = 2000
            content_length = len(book.content)
            return max(1, content_length // CHARS_PER_PAGE)
        
        # در غیر این صورت مقدار پیش‌فرض
        return 100

    @app.route('/user_dashboard')
    @login_required
    def user_dashboard():
        # داده‌های اولیه برای کارت‌ها و نمودارها
        total_users = User.query.count()
        total_books = Book.query.count()
        total_reads = ReadingProgress.query.count()

        # Get user's currently reading books (not completed)
        progresses_query = (ReadingProgress.query
                          .filter_by(user_id=current_user.id, completed=False)
                          .join(Book, ReadingProgress.book_id == Book.id)
                          .order_by(ReadingProgress.last_opened.desc()))
        # Get user's completed books
        completed_books = (ReadingProgress.query
                        .filter_by(user_id=current_user.id, completed=True)
                        .join(Book, ReadingProgress.book_id == Book.id)
                        .order_by(ReadingProgress.last_opened.desc())
                        .all())
        
        # Add page information to each progress
        progresses = []
        for progress in progresses_query.all():
            # Add page information to progress object
            progress.current_page = progress.position
            progress.total_pages = estimate_total_pages(progress.book) if progress.book else 100
            progress.percentage = min(100, int((progress.current_page / progress.total_pages) * 100)) if progress.total_pages > 0 else 0
            progresses.append(progress)
            
        # Get user's completed books
        completed_books_query = (ReadingProgress.query
                               .filter_by(user_id=current_user.id, completed=True)
                               .join(Book, ReadingProgress.book_id == Book.id)
                               .order_by(ReadingProgress.last_opened.desc()))
        
        # Add book information to completed books
        completed_books = []
        for progress in completed_books_query.all():
            if progress.book:  # Only include if book exists
                completed_books.append(progress.book)

        top_books = Book.query.order_by(Book.read_count.desc()).limit(5).all()

        # books per user
        books_per_user = db.session.query(
            User.username,
            func.count(ReadingProgress.id).label('books_read')
        ).outerjoin(ReadingProgress, User.id == ReadingProgress.user_id
        ).group_by(User.id).order_by(func.count(ReadingProgress.id).desc()).limit(10).all()

        # Get books being read by multiple users (currently reading)
        currently_reading_books = db.session.query(
            Book.id,
            Book.title,
            Book.author,
            Book.cover_filename,
            func.count(ReadingProgress.user_id).label('readers_count'),
            func.max(Book.read_count).label('total_reads')
        ).join(ReadingProgress, Book.id == ReadingProgress.book_id
        ).group_by(Book.id
        ).having(func.count(ReadingProgress.user_id) > 1
        ).order_by(func.count(ReadingProgress.user_id).desc()
        ).limit(5).all()

        # Get books completed by multiple users
        popular_completed_books = db.session.query(
            Book.id,
            Book.title,
            Book.author,
            Book.cover_filename,
            func.count(ReadingProgress.user_id).label('completed_count')
        ).join(ReadingProgress, Book.id == ReadingProgress.book_id
        ).group_by(Book.id
        ).having(func.count(ReadingProgress.user_id) > 1
        ).order_by(func.count(ReadingProgress.user_id).desc()
        ).limit(5).all()

        # prepare chart data
        book_titles = [b.title for b in top_books]
        book_reads = [b.read_count for b in top_books]

        return render_template('user_dashboard.html',
                               users=total_users,
                               books=total_books,
                               total_reads=total_reads,
                               top_books=zip(book_titles, book_reads),
                               books_per_user=books_per_user,
                               progresses=progresses,
                               currently_reading_books=currently_reading_books,
                               completed_books=completed_books)

    # ---------- ثبت‌نام و ورود ----------
    @app.route('/register', methods=['GET', 'POST'])
    @limiter.limit('10 per hour', methods=['POST'])
    def register():
        if request.method == 'POST':
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '')
            recovery = request.form.get('recovery', '').strip()
            email = request.form.get('email', '').strip().lower()
            if not username or not password or not recovery or not email:
                flash('لطفاً همهٔ فیلدها (از جمله ایمیل) را پر کنید.', 'danger')
                return redirect(url_for('register'))
            if not EMAIL_RE.match(email):
                flash('ایمیل وارد شده معتبر نیست.', 'danger')
                return redirect(url_for('register'))
            if User.query.filter_by(username=username).first():
                flash('این نام کاربری قبلاً ثبت شده.', 'danger')
                return redirect(url_for('register'))
            if User.query.filter_by(email=email).first():
                flash('این ایمیل قبلاً برای حساب دیگری ثبت شده.', 'danger')
                return redirect(url_for('register'))
            # هش کردن رمز عبور با Bcrypt
            hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
            hashed_recovery = bcrypt.generate_password_hash(recovery).decode('utf-8')
            user = User(username=username,
                        password_hash=hashed_password,
                        recovery_code=hashed_recovery,
                        email=email,
                        email_verified=False)
            db.session.add(user)
            db.session.commit()
            email_sent = send_verification_email(app, user)
            if email_sent:
                flash(
                    Markup(
                        '🎉 ثبت‌نام شما با موفقیت انجام شد! یک ایمیل تأیید برای شما ارسال کردیم؛ '
                        'لطفاً صندوق ورودی خود را بررسی کنید و روی لینک تأیید کلیک کنید. '
                        '📥 اگر ایمیل را ندیدید، نگران نباشید — گاهی ایمیل‌ها به‌اشتباه در پوشهٔ '
                        '<strong>«اسپم» (Spam) یا «تبلیغات» (Promotions)</strong> قرار می‌گیرند؛ '
                        'حتماً آنجا را هم چک کنید.'
                    ),
                    'success'
                )
            else:
                # ثبت‌نام موفق بوده ولی ارسال ایمیل شکست خورده (مثلاً تنظیمات
                # SMTP اشتباهه) — کاربر رو بی‌خبر نمی‌ذاریم و راه جایگزین
                # (ارسال دوباره) رو بهش نشون می‌دیم.
                flash(
                    Markup(
                        'ثبت‌نام شما با موفقیت انجام شد، اما در ارسال ایمیل تأیید مشکلی پیش آمد. '
                        f'می‌توانید از صفحهٔ <a href="{url_for("resend_verification")}">«ارسال دوباره ایمیل تأیید»</a> '
                        'دوباره تلاش کنید.'
                    ),
                    'warning'
                )
            return redirect(url_for('login'))
        return render_template('register.html')

    @app.route('/verify_email/<token>')
    def verify_email(token):
        email = verify_token(app, token, EMAIL_VERIFY_SALT, max_age_seconds=60 * 60 * 24)
        if not email:
            flash('لینک تأیید ایمیل نامعتبر یا منقضی‌شده است.', 'danger')
            return redirect(url_for('login'))
        user = User.query.filter_by(email=email).first()
        if not user:
            flash('کاربری با این ایمیل پیدا نشد.', 'danger')
            return redirect(url_for('login'))
        if not user.email_verified:
            user.email_verified = True
            db.session.commit()
        flash('ایمیل شما با موفقیت تأیید شد.', 'success')
        return redirect(url_for('login'))

    @app.route('/resend_verification', methods=['GET', 'POST'])
    @limiter.limit('5 per hour', methods=['POST'])
    def resend_verification():
        if request.method == 'POST':
            username = request.form.get('username', '').strip()
            user = User.query.filter_by(username=username).first()
            # پیام یکسان چه کاربر پیدا بشه چه نشه، تا کسی نتونه با این فرم
            # حدس بزنه چه نام‌کاربری‌هایی در سیستم وجود داره.
            if user and user.email and not user.email_verified:
                send_verification_email(app, user)
            flash('اگر حساب با این نام‌کاربری وجود داشته باشد و ایمیل تأیید نشده باشد، ایمیل تأیید دوباره ارسال شد.', 'info')
            return redirect(url_for('login'))
        return render_template('resend_verification.html')

    @app.route('/login', methods=['GET', 'POST'])
    @limiter.limit('10 per minute', methods=['POST'])
    def login():
        if request.method == 'POST':
            username_input = request.form.get('username') or ''
            user = User.query.filter_by(username=username_input).first()
            if user and bcrypt.check_password_hash(user.password_hash, request.form.get('password') or ''):
                # اگه کاربر ایمیل ثبت کرده ولی هنوز تأییدش نکرده، اجازهٔ ورود
                # نده. (حساب‌های قدیمی که اصلاً ایمیل ندارن مستثنا هستن تا
                # قفل نشن.)
                if user.email and not user.email_verified:
                    flash(
                        Markup(
                            'ایمیل شما هنوز تأیید نشده است. برای ورود، ابتدا باید روی لینک '
                            'ارسال‌شده به ایمیل خود کلیک کنید. اگر ایمیل را پیدا نمی‌کنید، '
                            'پوشهٔ اسپم را چک کنید یا از '
                            f'<a href="{url_for("resend_verification")}">اینجا</a> دوباره درخواست ارسال بدهید.'
                        ),
                        'warning'
                    )
                    return render_template('login.html')
                # Update last_login time
                user.last_login = datetime.utcnow()
                db.session.add(ActivityLog(user_id=user.id, action='login'))
                db.session.commit()
                login_user(user)
                next_page = request.args.get('next')
                if next_page and is_safe_url(next_page):
                    return redirect(next_page)
                return redirect(url_for('index'))
            flash('نام کاربری یا رمز عبور نامعتبر است', 'error')
        return render_template('login.html')

    @app.route('/forgot', methods=['GET', 'POST'])
    @limiter.limit('10 per hour', methods=['POST'])
    def forgot():
        # مسیر اصلی بازیابی رمز عبور: با ایمیل. یک لینک یک‌بارمصرف (معتبر تا
        # ۱ ساعت) به ایمیل ثبت‌شده‌ی کاربر ارسال می‌شه.
        if request.method == 'POST':
            email = request.form.get('email', '').strip().lower()
            if not email or not EMAIL_RE.match(email):
                flash('لطفاً یک ایمیل معتبر وارد کنید.', 'danger')
                return redirect(url_for('forgot'))
            user = User.query.filter_by(email=email).first()
            if user:
                send_password_reset_email(app, user)
            # پیام یکسان چه ایمیل پیدا بشه چه نشه، تا کسی نتونه با این فرم
            # حدس بزنه چه ایمیل‌هایی در سیستم ثبت شده.
            flash('اگر این ایمیل در سیستم ثبت شده باشد، لینک بازیابی رمز عبور برای آن ارسال شد.', 'info')
            return redirect(url_for('login'))
        return render_template('forgot.html')

    @app.route('/forgot_recovery', methods=['GET', 'POST'])
    @limiter.limit('10 per hour', methods=['POST'])
    def forgot_recovery():
        # مسیر جایگزین (قدیمی) برای کاربرانی که هنوز ایمیلی برای حساب خود ثبت
        # نکرده‌اند و بنابراین نمی‌توانند از مسیر اصلی /forgot استفاده کنند.
        if request.method == 'POST':
            username = request.form.get('username', '').strip()
            recovery_code = request.form.get('recovery', '').strip()
            user = User.query.filter_by(username=username).first()
            if user and bcrypt.check_password_hash(user.recovery_code, recovery_code):
                new_password = request.form.get('new_password', '')
                confirm_password = request.form.get('confirm_password', '')
                if not new_password or len(new_password) < 8:
                    flash('رمز عبور باید حداقل ۸ کاراکتر باشد.', 'danger')
                    return redirect(url_for('forgot_recovery'))
                if new_password != confirm_password:
                    flash('رمز عبور و تکرار آن یکسان نیستند.', 'danger')
                    return redirect(url_for('forgot_recovery'))
                user.password_hash = bcrypt.generate_password_hash(new_password).decode('utf-8')
                db.session.commit()
                flash('رمز عبور با موفقیت تغییر کرد.', 'success')
                return redirect(url_for('login'))
            flash('نام‌کاربری یا کد بازیابی اشتباه است.', 'danger')
            return redirect(url_for('forgot_recovery'))
        return render_template('forgot_recovery.html')

    @app.route('/reset_password/<token>', methods=['GET', 'POST'])
    @limiter.limit('10 per hour', methods=['POST'])
    def reset_password(token):
        email = verify_token(app, token, PASSWORD_RESET_SALT, max_age_seconds=60 * 60)
        if not email:
            flash('لینک بازیابی رمز عبور نامعتبر یا منقضی‌شده است.', 'danger')
            return redirect(url_for('forgot'))
        user = User.query.filter_by(email=email).first()
        if not user:
            flash('کاربری با این ایمیل پیدا نشد.', 'danger')
            return redirect(url_for('forgot'))
        if request.method == 'POST':
            new_password = request.form.get('new_password', '')
            confirm_password = request.form.get('confirm_password', '')
            if not new_password or len(new_password) < 8:
                flash('رمز عبور باید حداقل ۸ کاراکتر باشد.', 'danger')
                return redirect(url_for('reset_password', token=token))
            if new_password != confirm_password:
                flash('رمز عبور و تکرار آن یکسان نیستند.', 'danger')
                return redirect(url_for('reset_password', token=token))
            user.password_hash = bcrypt.generate_password_hash(new_password).decode('utf-8')
            db.session.commit()
            flash('رمز عبور با موفقیت تغییر کرد. اکنون می‌توانید وارد شوید.', 'success')
            return redirect(url_for('login'))
        return render_template('reset_password.html', token=token)

    @app.route('/logout')
    @login_required
    def logout():
        db.session.add(ActivityLog(user_id=current_user.id, action='logout'))
        db.session.commit()
        logout_user()
        flash('خروج انجام شد.', 'info')
        return redirect(url_for('login'))

    # ---------- لیست کتاب‌ها ----------
    @app.route('/books')
    @login_required
    def books():
        q = request.args.get('q','').strip()
        category = request.args.get('category','')
        page = request.args.get('page', 1, type=int)
        per_page = 8
        query = Book.query
        if q:
            query = query.filter((Book.title.contains(q)) | (Book.author.contains(q)))
        if category:
            query = query.filter_by(category=category)
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        list_books = pagination.items
        return render_template('book_list.html', books=list_books, q=q, selected_category=category, pagination=pagination)

        # ---------- خواندن کتاب ----------
    @app.route('/read/<int:book_id>', methods=['GET', 'POST'])
    @login_required
    def read_book(book_id):
        book = Book.query.get_or_404(book_id)
        if not book.file_type or book.file_type.lower() != 'pdf':
            flash('این کتاب قابل نمایش نیست', 'error')
            return redirect(url_for('books'))
            
        current_user.last_read_book_id = book_id
        
        # Get or create reading progress
        progress = ReadingProgress.query.filter_by(
            user_id=current_user.id, 
            book_id=book_id
        ).first()
        
        if not progress:
            progress = ReadingProgress(
                user_id=current_user.id, 
                book_id=book_id, 
                position=0,
                last_opened=datetime.utcnow()
            )
            db.session.add(progress)
        else:
            progress.last_opened = datetime.utcnow()
        
        db.session.commit()
        
        return redirect(url_for('read_pdf', book_id=book_id))

    # ---------- خواندن PDF ----------
    @app.route('/read_pdf/<int:book_id>')
    @login_required
    def read_pdf(book_id):
        book = Book.query.get_or_404(book_id)
        if not book.file_type or book.file_type.lower() != 'pdf':
            flash("این کتاب PDF نیست.", "danger")
            return redirect(url_for('books'))

        if not user_has_access(book):
            flash('برای مطالعه‌ی این کتاب باید ابتدا آن را خریداری کنید.', 'warning')
            return redirect(url_for('buy_book', book_id=book.id))
        
        # Get or create reading progress
        progress = ReadingProgress.query.filter_by(
            user_id=current_user.id, 
            book_id=book.id
        ).first()
        
        if not progress:
            progress = ReadingProgress(
                user_id=current_user.id,
                book_id=book.id,
                position=0,  # For PDFs, position will be the page number
                last_opened=datetime.utcnow()
            )
            db.session.add(progress)
        else:
            progress.last_opened = datetime.utcnow()
        
        db.session.commit()
        
        return render_template("read_pdf.html", 
                            book=book, 
                            progress=progress,
                            pdf_file=book.pdf_filename)
    
    @app.route('/get_book_file/<int:book_id>')
    @login_required
    def get_book_file(book_id):
        """Serve book files for the PDF.js viewer"""
        try:
            book = Book.query.get_or_404(book_id)

            if not user_has_access(book):
                return "برای مشاهده‌ی این کتاب باید ابتدا آن را خریداری کنید.", 403

            file_path = get_book_file_path(book)
            if not file_path:
                return "فرمت فایل پشتیبانی نمی‌شود", 400

            mime_type = get_mime_for_book_type(book.file_type)

            if not os.path.exists(file_path):
                app.logger.error(f"File not found: {file_path}")
                return "فایل یافت نشد", 404

            return send_file(
                file_path,
                mimetype=mime_type,
                as_attachment=False,
                download_name=book.pdf_filename
            )

        except Exception as e:
            app.logger.error(f"Error serving book file {book_id}: {str(e)}")
            return str(e), 500

    # ---------- خرید و پرداخت کتاب ----------

    @app.route('/book/<int:book_id>/buy')
    @login_required
    def buy_book(book_id):
        book = Book.query.get_or_404(book_id)
        if not book.price or book.price <= 0:
            flash('این کتاب رایگان است.', 'info')
            return redirect(url_for('view_book', book_id=book.id))
        if user_has_access(book):
            flash('شما قبلاً این کتاب را خریداری کرده‌اید.', 'info')
            return redirect(url_for('view_book', book_id=book.id))

        pending = Purchase.query.filter_by(
            user_id=current_user.id, book_id=book.id, status='pending'
        ).order_by(Purchase.created_at.desc()).first()

        return render_template('buy_book.html', book=book, pending=pending,
                                payment_mode=app.config['PAYMENT_MODE'])

    @app.route('/book/<int:book_id>/checkout', methods=['POST'])
    @login_required
    @limiter.limit('20 per hour')
    def checkout_book(book_id):
        book = Book.query.get_or_404(book_id)
        if not book.price or book.price <= 0 or user_has_access(book):
            return redirect(url_for('view_book', book_id=book.id))

        purchase = Purchase(
            user_id=current_user.id,
            book_id=book.id,
            amount=book.price,
            status='pending',
            gateway=app.config['PAYMENT_MODE']
        )
        db.session.add(purchase)
        db.session.commit()

        gateway = get_gateway(app.config)
        callback_url = url_for('payment_callback', purchase_id=purchase.id, _external=True)
        result = gateway.request_payment(
            amount=book.price,
            description=f"خرید کتاب: {book.title}",
            callback_url=callback_url,
            order_id=purchase.id
        )

        if not result.success:
            purchase.status = 'failed'
            db.session.commit()
            flash(f'اتصال به درگاه پرداخت ناموفق بود: {result.message}', 'error')
            return redirect(url_for('buy_book', book_id=book.id))

        purchase.track_id = result.track_id
        db.session.commit()

        if gateway.name == 'sandbox':
            return redirect(url_for('sandbox_payment_page', purchase_id=purchase.id))
        return redirect(result.payment_url)

    @app.route('/payment/sandbox/<int:purchase_id>', methods=['GET', 'POST'])
    @login_required
    def sandbox_payment_page(purchase_id):
        """صفحه‌ی شبیه‌سازی درگاه پرداخت. فقط وقتی PAYMENT_MODE == 'sandbox' فعاله."""
        if app.config['PAYMENT_MODE'] != 'sandbox':
            abort(404)
        purchase = Purchase.query.get_or_404(purchase_id)
        if purchase.user_id != current_user.id:
            abort(403)
        if purchase.status != 'pending':
            return redirect(url_for('my_purchases'))

        if request.method == 'POST':
            outcome = request.form.get('outcome')
            return redirect(url_for('payment_callback', purchase_id=purchase.id, outcome=outcome))

        return render_template('payment_sandbox.html', purchase=purchase)

    @app.route('/payment/callback/<int:purchase_id>')
    def payment_callback(purchase_id):
        purchase = Purchase.query.get_or_404(purchase_id)

        if purchase.status == 'paid':
            flash('این پرداخت قبلاً تایید شده است.', 'info')
            return redirect(url_for('my_purchases'))

        # دفاع عمیق‌تر: اگر کاربری session فعال دارد، حتماً باید مالک خرید باشد.
        # در حالت عادی کاربر از درگاه به همین صفحه هدایت می‌شود و session فعال است.
        if current_user.is_authenticated and purchase.user_id != current_user.id:
            app.logger.warning(
                f"کاربر {current_user.id} تلاش تایید پرداخت purchase={purchase_id} "
                f"(متعلق به کاربر {purchase.user_id}) را داشت."
            )
            abort(403)

        gateway = get_gateway(app.config)

        if gateway.name == 'sandbox':
            outcome = request.args.get('outcome')
            if outcome == 'success':
                result = gateway.verify_payment(purchase.track_id, purchase.amount)
            else:
                result = PaymentResult(success=False, message='پرداخت توسط کاربر لغو شد (حالت آزمایشی)')
        else:
            # پارامترهای بازگشتی درگاه زیبال: trackId و success (0 یا 1)
            track_id = request.args.get('trackId', purchase.track_id)
            success_flag = request.args.get('success')
            if success_flag == '0':
                result = PaymentResult(success=False, message='پرداخت توسط کاربر لغو شد')
            else:
                result = gateway.verify_payment(track_id, purchase.amount)

        if result.success:
            purchase.status = 'paid'
            purchase.ref_number = result.ref_number
            purchase.paid_at = datetime.utcnow()
            db.session.commit()
            flash(f'پرداخت با موفقیت انجام شد. کد پیگیری: {result.ref_number}', 'success')
            return redirect(url_for('view_book', book_id=purchase.book_id))
        else:
            purchase.status = 'failed'
            db.session.commit()
            flash(f'پرداخت ناموفق بود: {result.message}', 'error')
            return redirect(url_for('buy_book', book_id=purchase.book_id))

    @app.route('/my-purchases')
    @login_required
    def my_purchases():
        purchases = Purchase.query.filter_by(user_id=current_user.id, status='paid') \
            .order_by(Purchase.paid_at.desc()).all()
        return render_template('my_purchases.html', purchases=purchases)

    @app.route('/admin/sales')
    @login_required
    def admin_sales():
        if not current_user.is_admin:
            flash('شما مجاز به مشاهده این صفحه نیستید', 'danger')
            return redirect(url_for('books'))
        purchases = Purchase.query.filter_by(status='paid') \
            .order_by(Purchase.paid_at.desc()).all()
        total_revenue = sum(p.amount for p in purchases)
        return render_template('admin_sales.html', purchases=purchases, total_revenue=total_revenue)

    @app.route('/advanced-reports')
    @login_required
    @admin_required
    def admin_advanced_reports():
        graph_data = {
            "data": [{"x": [1, 2, 3], "y": [4, 5, 6], "type": "scatter"}],
            "layout": {"title": "نمودار من"}
        }
        return render_template('advanced_reports.html', graph_json=graph_data)

    @app.route('/view_book/<int:book_id>')
    @login_required
    def view_book(book_id):
        try:
            book = Book.query.get_or_404(book_id)

            if not user_has_access(book):
                flash('برای مطالعه‌ی این کتاب باید ابتدا آن را خریداری کنید.', 'warning')
                return redirect(url_for('buy_book', book_id=book.id))

            # Verify the book has a file
            if not book.pdf_filename:
                flash('فایل کتاب یافت نشد', 'error')
                return redirect(url_for('books'))
            
            # Track reading progress
            progress = db.session.query(ReadingProgress).filter(
                ReadingProgress.user_id == current_user.id,
                ReadingProgress.book_id == book_id
            ).first()
            
            if not progress:
                progress = ReadingProgress(
                    user_id=current_user.id,
                    book_id=book_id,
                    position=1,
                    last_opened=datetime.utcnow(),
                    total_pages=1  # Initialize with a default value
                )
                db.session.add(progress)
                db.session.commit()
            else:
                progress.last_opened = datetime.utcnow()
                db.session.commit()
            
            file_path = get_book_file_path(book)
            if not file_path:
                flash('فرمت فایل پشتیبانی نمی‌شود', 'error')
                return redirect(url_for('books'))
            app.logger.info(f"Looking for book file at: {file_path} (exists={os.path.exists(file_path)})")
                
            # Verify the file exists
            if not os.path.exists(file_path):
                app.logger.error(f"File not found at: {file_path}")
                flash('فایل مورد نظر یافت نشد', 'error')
                return redirect(url_for('books'))
            
            # Handle PDF files
            if book.file_type.lower() == 'pdf':
                # Verify it's a valid PDF
                try:
                    with open(file_path, 'rb') as f:
                        header = f.read(4)
                        if header != b'%PDF':
                            app.logger.error(f"Invalid PDF header in file: {file_path}")
                            flash('فایل PDF معتبر نیست', 'error')
                            return redirect(url_for('books'))
                except Exception as e:
                    app.logger.error(f"Error reading PDF file {file_path}: {str(e)}")
                    flash('خطا در خواندن فایل PDF', 'error')
                    return redirect(url_for('books'))
                
                # Update read count
                book.read_count += 1
                db.session.commit()
                
                # Convert progress to a dictionary to avoid SQLAlchemy lazy loading issues
                progress_data = {
                    'id': progress.id,
                    'user_id': progress.user_id,
                    'book_id': progress.book_id,
                    'position': progress.position,
                    'total_pages': progress.total_pages,
                    'last_opened': progress.last_opened
                }
                
                return render_template('pdf_viewer.html',
                                    book=book,
                                    progress=progress_data,
                                    file_path=book.pdf_filename)
            
            # Handle TXT files
            elif book.file_type.lower() == 'txt':
                try:
                    # Read the entire content
                    with open(file_path, 'r', encoding='utf-8') as f:
                        full_content = f.read()
                    
                    # Split content into pages (approximately 2000 characters per page)
                    page_size = 2000
                    pages = [full_content[i:i+page_size] for i in range(0, len(full_content), page_size)]
                    total_pages = len(pages) or 1
                    
                    # Get current page from query parameter or progress
                    current_page = request.args.get('page', type=int) or (progress.position if progress and progress.position > 0 else 1)
                    current_page = max(1, min(current_page, total_pages))
                    
                    # Get content for current page
                    current_content = pages[current_page - 1] if pages else ''
                    
                    # Ensure progress exists
                    if not progress:
                        progress = ReadingProgress(
                            user_id=current_user.id,
                            book_id=book_id,
                            position=current_page,
                            total_pages=total_pages,
                            last_opened=datetime.utcnow()
                        )
                        db.session.add(progress)
                    # Update progress if needed
                    elif progress.position != current_page or not hasattr(progress, 'total_pages') or progress.total_pages != total_pages:
                        progress.position = current_page
                        progress.total_pages = total_pages
                        progress.last_opened = datetime.utcnow()
                    
                    db.session.commit()
                    
                    # Update read count if this is the first time viewing this page
                    if current_page == 1 and not request.args.get('page'):
                        book.read_count += 1
                        db.session.commit()
                    
                    # Convert progress to a dictionary to avoid SQLAlchemy lazy loading issues
                    progress_data = {
                        'id': progress.id,
                        'user_id': progress.user_id,
                        'book_id': progress.book_id,
                        'position': progress.position,
                        'total_pages': progress.total_pages,
                        'last_opened': progress.last_opened
                    }
                    
                    return render_template('text_viewer.html',
                                        book=book,
                                        progress=progress_data,
                                        content=current_content,
                                        current_page=current_page,
                                        total_pages=total_pages,
                                        full_content_length=len(full_content))
                                        
                except Exception as e:
                    app.logger.error(f"Error reading TXT file {file_path}: {str(e)}")
                    flash('خطا در خواندن فایل متنی', 'error')
                    return redirect(url_for('books'))
                    
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error in view_book for book {book_id}: {str(e)}", exc_info=True)
            flash('خطا در نمایش کتاب', 'error')
            return redirect(url_for('books'))
            
    def _save_progress_logic(user_id, book_id, position, total_pages=None, mark_completed=False):
        """منطق مشترک ذخیره پیشرفت، برای استفاده از هر دو مسیر form و JSON."""
        progress = ReadingProgress.query.filter_by(
            user_id=user_id,
            book_id=book_id
        ).first()

        if not progress:
            progress = ReadingProgress(
                user_id=user_id,
                book_id=book_id,
                position=position,
                total_pages=total_pages or 1,
                last_opened=datetime.utcnow(),
                completed=mark_completed
            )
            db.session.add(progress)
        else:
            progress.position = position
            if total_pages and total_pages > 0:
                progress.total_pages = total_pages
            if mark_completed:
                progress.completed = True
            progress.last_opened = datetime.utcnow()

        db.session.commit()
        return progress

    @app.route('/save_text_progress/<int:book_id>', methods=['POST'])
    @login_required
    def save_text_progress(book_id):
        return save_reading_progress(book_id)

    @app.route('/save_reading_progress/<int:book_id>', methods=['POST'])
    @login_required
    def save_reading_progress(book_id):
        try:
            position = request.form.get('position', type=int) or 1
            total_pages = request.form.get('total_pages', type=int)
            mark_completed = request.form.get('mark_completed', 'false').lower() == 'true'

            progress = _save_progress_logic(
                user_id=current_user.id,
                book_id=book_id,
                position=position,
                total_pages=total_pages,
                mark_completed=mark_completed,
            )
            return jsonify({
                'success': True,
                'position': progress.position,
                'total_pages': progress.total_pages,
                'completed': progress.completed
            })
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error saving progress: {str(e)}")
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/api/save_reading_progress', methods=['POST'])
    @login_required
    def api_save_reading_progress():
        """ذخیره پیشرفت با JSON body (هماهنگ با pdf_viewer.js)."""
        try:
            data = request.get_json(silent=True) or {}
            book_id = data.get('book_id')
            position = int(data.get('position', 1) or 1)
            total_pages = data.get('total_pages')
            if total_pages is not None:
                total_pages = int(total_pages)
            mark_completed = str(data.get('mark_completed', 'false')).lower() == 'true'

            if not book_id:
                return jsonify({'success': False, 'error': 'book_id الزامی است'}), 400

            progress = _save_progress_logic(
                user_id=current_user.id,
                book_id=int(book_id),
                position=position,
                total_pages=total_pages,
                mark_completed=mark_completed,
            )
            return jsonify({
                'success': True,
                'position': progress.position,
                'total_pages': progress.total_pages,
                'completed': progress.completed
            })
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error in api_save_reading_progress: {str(e)}")
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/api/get_reading_progress')
    @login_required
    def api_get_reading_progress():
        """دریافت پیشرفت ذخیره‌شده یک کتاب برای کاربر جاری."""
        try:
            book_id = request.args.get('book_id', type=int)
            if not book_id:
                return jsonify({'success': False, 'error': 'book_id الزامی است'}), 400

            Book.query.get_or_404(book_id)

            progress = ReadingProgress.query.filter_by(
                user_id=current_user.id,
                book_id=book_id
            ).first()

            if not progress:
                return jsonify({
                    'success': True,
                    'book_id': book_id,
                    'position': 0,
                    'total_pages': 0,
                    'completed': False
                })

            return jsonify({
                'success': True,
                'book_id': book_id,
                'position': progress.position,
                'total_pages': progress.total_pages,
                'completed': progress.completed,
                'last_opened': progress.last_opened.isoformat() if progress.last_opened else None
            })
        except Exception as e:
            app.logger.error(f"Error in api_get_reading_progress: {str(e)}")
            return jsonify({'success': False, 'error': str(e)}), 500
            
    # In app.py, update the add_bookmark route:
    @app.route('/add_bookmark/<int:book_id>', methods=['POST'])
    @login_required
    def add_bookmark(book_id):
        try:
            page = request.form.get('page', type=int) or 1
            note = request.form.get('note', '')
            
            # Check if bookmark already exists
            existing = Bookmark.query.filter_by(
                user_id=current_user.id,
                book_id=book_id,
                page=page
            ).first()
            
            if existing:
                return jsonify({
                    'success': False,
                    'message': 'نشانک برای این صفحه قبلاً ذخیره شده است'
                })
                
            bookmark = Bookmark(
                user_id=current_user.id,
                book_id=book_id,
                page=page,
                note=note
            )
            db.session.add(bookmark)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'نشانک با موفقیت ذخیره شد'
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({
                'success': False,
                'message': f'خطا در ذخیره نشانک: {str(e)}'
            }), 500
            
    @app.route('/remove_bookmark/<int:bookmark_id>', methods=['DELETE'])
    @login_required
    def remove_bookmark(bookmark_id):
        try:
            bookmark = Bookmark.query.get_or_404(bookmark_id)
            
            # Verify ownership
            if bookmark.user_id != current_user.id:
                return jsonify({'success': False, 'error': 'دسترسی غیرمجاز'}), 403
            
            db.session.delete(bookmark)
            db.session.commit()
            
            return jsonify({'success': True})
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error removing bookmark: {str(e)}")
            return jsonify({'success': False, 'error': str(e)}), 500
            
    @app.route('/get_bookmarks/<int:book_id>')
    @login_required
    def get_bookmarks_for_book(book_id):
        try:
            # Get the book
            book = Book.query.get_or_404(book_id)
            
            # Get all bookmarks for this book and user
            bookmarks = Bookmark.query.filter_by(
                user_id=current_user.id,
                book_id=book_id
            ).order_by(Bookmark.page.asc()).all()
            
            # Get current page from query params
            current_page = int(request.args.get('current_page', 0))
            
            # Convert bookmarks to a list of dictionaries
            bookmarks_data = [{
                'id': b.id,
                'page': b.page,
                'note': b.note,
                'created_at': b.created_at.isoformat() if b.created_at else None,
                'is_current': b.page == current_page
            } for b in bookmarks]
            
            return jsonify({
                'success': True,
                'bookmarks': bookmarks_data,
                'book': {
                    'id': book.id,
                    'title': book.title,
                    'file_type': book.file_type
                }
            })
            
        except Exception as e:
            app.logger.error(f"Error getting bookmarks: {str(e)}")
            return jsonify({
                'success': False, 
                'error': 'خطا در دریافت نشانک‌ها',
                'details': str(e)
            }), 500

    @app.route('/mark_books_seen', methods=['POST'])
    @login_required
    def mark_books_seen():
        try:
            current_user.last_checked_books = datetime.utcnow()
            db.session.commit()
            return '', 204
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error in mark_books_seen: {str(e)}")
            return jsonify({'error': 'خطا در بروزرسانی'}), 500

    @app.route('/new_books')
    @login_required
    def new_books():
        # Get new books since last check
        last_checked = _strip_tz(current_user.last_checked_books)
        new_books_list = Book.query.filter(Book.created_at > last_checked).all()

        # Update the last checked time
        current_user.last_checked_books = datetime.utcnow()
        db.session.commit()

        return render_template('new_books.html', new_books=new_books_list)


    # ---------- نشانک ----------
    # Bookmark routes
    @app.route('/toggle_bookmark/<int:book_id>', methods=['POST'])
    @login_required
    def toggle_bookmark(book_id):
        try:
            book = Book.query.get_or_404(book_id)
            data = request.get_json() or {}
            
            # Get the page number and note from the request
            page = int(data.get('page', 1))
            note = data.get('note', '')
            action = data.get('action', 'add')  # 'add' or 'remove'

            # Check if the book exists
            book = Book.query.get_or_404(book_id)

            # If bookmark_id is provided, use it to find the bookmark
            bookmark_id = data.get('bookmark_id')
            if action == 'remove' and bookmark_id:
                bookmark = Bookmark.query.filter_by(
                    id=bookmark_id,
                    user_id=current_user.id,
                    book_id=book_id
                ).first()
            else:
                # Otherwise, find by book_id and page
                bookmark = Bookmark.query.filter_by(
                    user_id=current_user.id,
                    book_id=book_id,
                    page=page
                ).first()

            if action == 'add':
                if bookmark:
                    # Update existing bookmark
                    bookmark.note = note
                    bookmark.created_at = datetime.utcnow()
                    message = 'نشانک به‌روزرسانی شد'
                else:
                    # Create new bookmark
                    bookmark = Bookmark(
                        user_id=current_user.id,
                        book_id=book_id,
                        page=page,
                        note=note
                    )
                    db.session.add(bookmark)
                    message = 'صفحه با موفقیت نشانک‌گذاری شد'
                
                db.session.commit()
                return jsonify({
                    'success': True,
                    'message': message,
                    'bookmark_id': bookmark.id,
                    'is_bookmarked': True,
                    'page': page,
                    'note': note
                })
                
            elif action == 'remove':
                if not bookmark:
                    return jsonify({
                        'success': False,
                        'error': 'نشانک یافت نشد'
                    }), 404
                    
                db.session.delete(bookmark)
                db.session.commit()
                return jsonify({
                    'success': True,
                    'message': 'نشانک حذف شد',
                    'is_bookmarked': False,
                    'page': page
                })
                
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Error in toggle_bookmark: {str(e)}')
            return jsonify({
                'success': False,
                'error': 'خطا در پردازش درخواست نشانک',
                'details': str(e)
            }), 500

    @app.route('/get_bookmarks')
    @login_required
    def get_all_bookmarks():
        try:
            # Get the latest bookmark for each book
            latest_bookmarks = db.session.query(
                Bookmark.book_id,
                db.func.max(Bookmark.created_at).label('latest_created_at')
            ).filter(
                Bookmark.user_id == current_user.id
            ).group_by(Bookmark.book_id).subquery()

            # Get book details with their latest bookmark
            bookmarks = db.session.query(Bookmark, Book).join(
                Book, Bookmark.book_id == Book.id
            ).join(
                latest_bookmarks,
                db.and_(
                    Bookmark.book_id == latest_bookmarks.c.book_id,
                    Bookmark.created_at == latest_bookmarks.c.latest_created_at
                )
            ).filter(
                Bookmark.user_id == current_user.id
            ).order_by(Bookmark.created_at.desc()).all()
            
            book_list = []
            for bm, book in bookmarks:
                book_list.append({
                    'id': book.id,
                    'title': book.title,
                    'author': book.author,
                    'latest_page': bm.page,  # Latest bookmarked page
                    'bookmark_count': Bookmark.query.filter_by(
                        user_id=current_user.id,
                        book_id=book.id
                    ).count(),  # Total bookmarks for this book
                    'cover_url': url_for('show_cover', filename=book.cover_filename) if book.cover_filename else None,
                    'last_bookmarked': bm.created_at.strftime('%Y-%m-%d') if bm.created_at else '',
                    'category': book.category
                })
            return jsonify(book_list)
        except Exception as e:
            app.logger.error(f"Error getting bookmarks: {str(e)}")
            return jsonify({'error': 'خطا در دریافت نشانک‌ها'}), 500
            
    # File serving routes
    @app.route('/serve_pdf/<path:filename>')
    @login_required
    def serve_pdf(filename):
        try:
            from urllib.parse import unquote
            filename = unquote(filename)

            if '..' in filename or filename.startswith('/'):
                flash('مسیر فایل نامعتبر است', 'error')
                return redirect(url_for('books'))

            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

            if not os.path.exists(filepath):
                ft_dirs = ['pdf', 'txt', 'epub']
                for d in ft_dirs:
                    alt = os.path.join(app.config['UPLOAD_FOLDER'], 'books', d, filename)
                    if os.path.exists(alt):
                        filepath = alt
                        break

            expected_upload_dir = os.path.abspath(app.config['UPLOAD_FOLDER'])
            resolved_path = os.path.abspath(filepath)
            if not resolved_path.startswith(expected_upload_dir + os.sep):
                app.logger.warning(f"تلاش برای دسترسی خارج از پوشه‌ی مجاز: {filename}")
                flash('مسیر فایل نامعتبر است', 'error')
                return redirect(url_for('books'))

            if not os.path.exists(filepath):
                app.logger.error(f"فایل یافت نشد: {filename}")
                flash('فایل مورد نظر یافت نشد', 'error')
                return redirect(url_for('books'))

            return send_file(
                filepath,
                as_attachment=False,
                download_name=os.path.basename(filename),
                mimetype='application/pdf'
            )

        except Exception as e:
            app.logger.error(f"خطا در سرویس‌دهی فایل: {str(e)}", exc_info=True)
            flash('خطا در نمایش فایل', 'error')
            return redirect(url_for('books'))

    @app.route('/cover/<path:filename>')
    @login_required
    def show_cover(filename):
        try:
            # Decode URL-encoded filename
            from urllib.parse import unquote
            filename = unquote(filename)
            
            # Use the UPLOAD_FOLDER/covers path for cover images
            cover_path = os.path.join(app.config['UPLOAD_FOLDER'], 'covers', filename)
            
            # Check if the file exists and is within the allowed directory
            expected_dir = os.path.abspath(os.path.join(app.config['UPLOAD_FOLDER'], 'covers'))
            if os.path.isfile(cover_path) and os.path.dirname(os.path.abspath(cover_path)) == expected_dir:
                return send_file(cover_path)
                
        except Exception as e:
            app.logger.error(f"Error showing cover {filename}: {str(e)}")
            
        # Return default cover if there's an error or file not found
        return send_from_directory('static', 'images/default-cover.jpg')

    @app.route('/download_book/<int:book_id>')
    @login_required
    def download_book(book_id):
        book = Book.query.get_or_404(book_id)

        if not user_has_access(book):
            flash('برای دانلود این کتاب باید ابتدا آن را خریداری کنید.', 'warning')
            return redirect(url_for('buy_book', book_id=book.id))

        if not book.pdf_filename:
            flash('فایل کتاب یافت نشد', 'error')
            return redirect(url_for('books'))
            
        try:
            file_path = get_book_file_path(book)
            if not file_path:
                flash('فرمت فایل پشتیبانی نمی‌شود', 'error')
                return redirect(url_for('books'))

            mime_type = get_mime_for_book_type(book.file_type)

            if not os.path.exists(file_path):
                flash('فایل مورد نظر یافت نشد', 'error')
                return redirect(url_for('books'))

            safe_title = re.sub(r'[^\w\s-]', '', book.title).strip()
            safe_title = re.sub(r'[-\s]+', '-', safe_title)
            download_filename = f"{safe_title}.{book.file_type}"

            return send_file(
                file_path,
                as_attachment=True,
                download_name=download_filename,
                mimetype=mime_type
            )
            
        except Exception as e:
            app.logger.error(f"Error downloading book {book_id}: {str(e)}")
            flash('خطا در دانلود فایل', 'error')
            return redirect(url_for('books'))

    @app.route('/edit_book/<int:book_id>', methods=['GET', 'POST'])
    @login_required
    def edit_book(book_id):
        if not current_user.is_admin:
            flash('شما مجاز به انجام این کار نیستید', 'danger')
            return redirect(url_for('books'))
            
        book = Book.query.get_or_404(book_id)

        form = EditBookForm(obj=book)

        if form.validate_on_submit():
            try:
                book.title = form.title.data
                book.author = form.author.data
                book.category = form.category.data
                book.price = form.price.data or 0

                if form.cover.data and getattr(form.cover.data, 'filename', None):
                    if book.cover_filename and os.path.exists(os.path.join(app.config['COVER_FOLDER'], book.cover_filename)):
                        try:
                            os.remove(os.path.join(app.config['COVER_FOLDER'], book.cover_filename))
                        except Exception as e:
                            app.logger.error(f'Error removing old cover image: {str(e)}')

                    cover_file = form.cover.data
                    cover_ext = os.path.splitext(cover_file.filename)[1].lower()
                    cover_filename = f"{uuid.uuid4().hex}{cover_ext}"
                    cover_path = os.path.join(app.config['COVER_FOLDER'], cover_filename)

                    from PIL import Image
                    img = Image.open(cover_file)
                    img.thumbnail((300, 400))
                    img.save(cover_path)

                    book.cover_filename = cover_filename

                db.session.commit()
                flash('کتاب با موفقیت به‌روزرسانی شد.', 'success')
                return redirect(url_for('books'))
                
            except Exception as e:
                db.session.rollback()
                app.logger.error(f'Error updating book: {str(e)}')
                flash('خطا در به‌روزرسانی کتاب. لطفاً دوباره تلاش کنید.', 'danger')
        else:
            # Pre-populate form with existing data
            form.title.data = book.title
            form.author.data = book.author
            form.category.data = book.category
            form.price.data = book.price
        
        return render_template('edit_book.html', form=form, book=book)

    @app.route('/delete_book/<int:book_id>', methods=['POST'])
    @login_required
    def delete_book(book_id):
        if not current_user.is_admin:
            flash('فقط ادمین اجازه حذف دارد.', 'danger')
            return redirect(url_for('books'))
        book = Book.query.get_or_404(book_id)

        paths_to_check = []
        book_file_path = get_book_file_path(book)
        if book_file_path:
            paths_to_check.append((os.path.dirname(book_file_path), os.path.basename(book_file_path)))
        paths_to_check.append((app.config['COVER_FOLDER'], book.cover_filename))
        paths_to_check.append((os.path.join(app.config['BASE_DIR'], 'all_books', 'pdf'), book.pdf_filename))
        paths_to_check.append((os.path.join(app.config['BASE_DIR'], 'all_books', 'txt'), f"{book.title}.txt"))
        paths_to_check.append((os.path.join(app.config['BASE_DIR'], 'all_books', 'covers'), book.cover_filename))

        for dir_path, fname in paths_to_check:
            try:
                if fname and dir_path and os.path.exists(os.path.join(dir_path, fname)):
                    os.remove(os.path.join(dir_path, fname))
            except Exception:
                pass
        ReadingProgress.query.filter_by(book_id=book_id).delete()
        db.session.delete(book)
        db.session.commit()
        flash('کتاب با موفقیت حذف شد.', 'success')
        return redirect(url_for('books'))

    # ---------- Admin User Management Routes ----------
    @app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
    @login_required
    @admin_required
    def admin_delete_user(user_id):
        if not current_user.is_admin:
            flash('دسترسی غیر مجاز', 'danger')
            return redirect(url_for('admin_dashboard'))

        user = User.query.get_or_404(user_id)
        if user.is_admin:
            flash('امکان حذف ادمین وجود ندارد', 'warning')
            return redirect(url_for('admin_dashboard'))

        try:
            # Delete related records
            ReadingProgress.query.filter_by(user_id=user.id).delete()
            ActivityLog.query.filter_by(user_id=user.id).delete()
            
            # Delete the user
            db.session.delete(user)
            db.session.commit()
            
            flash(f'کاربر {user.username} با موفقیت حذف شد.', 'success')
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Error deleting user: {str(e)}')
            flash('خطا در حذف کاربر', 'danger')
            
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/reset_user/<int:user_id>', methods=['POST'])
    @login_required
    @admin_required
    def admin_reset_user(user_id):
        if not current_user.is_admin:
            flash('دسترسی غیر مجاز', 'danger')
            return redirect(url_for('admin_dashboard'))

        user = User.query.get_or_404(user_id)
        try:
            # به‌جای یک پسورد ثابت و قابل‌حدس (مثل 123456) که برای همه‌ی کاربران
            # یکسانه، یک پسورد موقت تصادفی تولید می‌کنیم و فقط یک‌بار به ادمین
            # نشون می‌دیم تا به کاربر منتقل کنه. بهتره کاربر بعد از ورود اول
            # مجبور به تغییر رمز بشه (این بخشش نیاز به فیلد اضافه در مدل داره).
            temp_password = uuid.uuid4().hex[:10]
            user.password_hash = bcrypt.generate_password_hash(temp_password).decode('utf-8')
            db.session.commit()
            flash(f'رمز عبور کاربر {user.username} به «{temp_password}» ریست شد. لطفاً این رمز را به کاربر اطلاع دهید و از او بخواهید آن را تغییر دهد.', 'success')
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Error resetting password: {str(e)}')
            flash('خطا در ریست کردن رمز عبور', 'danger')
            
        return redirect(url_for('admin_dashboard'))

    # ---------- مسیرهای مدیریتی ----------
    @app.route('/admin/dashboard')
    @login_required
    @admin_required
    def admin_dashboard():
        # Get filter parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        user_id = request.args.get('user_id')
        status = request.args.get('status', 'all')
        page = request.args.get('page', 1, type=int)
        per_page = 10

        # Base query for users
        users_query = User.query

        # Apply status filter
        if status == 'active':
            users_query = users_query.filter(User.last_login >= datetime.utcnow() - timedelta(days=30))
        elif status == 'inactive':
            users_query = users_query.filter(
                (User.last_login < datetime.utcnow() - timedelta(days=30)) | 
                (User.last_login == None)
            )

        # Apply user filter
        if user_id:
            users_query = users_query.filter(User.id == user_id)

        # Apply date range filter
        if start_date and end_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                end_date = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
                users_query = users_query.filter(
                    (User.last_login >= start_date) & 
                    (User.last_login <= end_date)
                )
            except ValueError:
                flash('فرمت تاریخ نامعتبر است', 'error')

        # Get paginated users
        users_pagination = users_query.order_by(User.username.asc()).paginate(
            page=page, 
            per_page=per_page,
            error_out=False
        )
        users_list = users_pagination.items

        # Get all users for the filter dropdown
        all_users = User.query.order_by(User.username).all()

        # Calculate statistics
        total_users = User.query.count()
        total_books = Book.query.count()
        total_reads = ReadingProgress.query.count()
        total_bookmarks = Bookmark.query.count()
        avg_books_read = round(float(total_reads) / float(total_users), 1) if total_users and total_users > 0 else 0.0

        # Get top books
        top_books = db.session.query(
            Book.title, 
            Book.author,
            func.count(ReadingProgress.id).label('read_count')
        ).join(ReadingProgress).group_by(Book.id).order_by(
            func.count(ReadingProgress.id).desc()
        ).limit(5).all()

        # Get books per user
        books_per_user = db.session.query(
            User.username, 
            func.count(ReadingProgress.id).label('books_read')
        ).outerjoin(ReadingProgress).group_by(User.id).order_by(
            func.count(ReadingProgress.id).desc()
        ).limit(10).all()

        # Get reading statistics
        currently_reading_stats = db.session.query(
            Book.title,
            func.count(ReadingProgress.user_id).label('readers_count')
        ).join(ReadingProgress).filter(
            ReadingProgress.completed == False
        ).group_by(Book.id).order_by(
            func.count(ReadingProgress.user_id).desc()
        ).limit(5).all()

        # Get completed books statistics
        completed_books_stats = db.session.query(
            Book.title,
            func.count(ReadingProgress.user_id).label('completed_count')
        ).join(ReadingProgress).filter(
            ReadingProgress.completed == True
        ).group_by(Book.id).order_by(
            func.count(ReadingProgress.user_id).desc()
        ).limit(5).all()

        # Get active readers
        active_readers = db.session.query(
            User.username,
            func.count(ReadingProgress.id).label('books_reading')
        ).join(ReadingProgress).group_by(User.id).order_by(
            func.count(ReadingProgress.id).desc()
        ).limit(10).all()

        # Calculate user statistics with progress
        user_stats = []
        for user in users_list:
            total_progress = []
            for progress in user.reading_progress.all():
                if progress.total_pages > 0:
                    progress_percentage = min(100, (progress.position / progress.total_pages) * 100)
                    total_progress.append(progress_percentage)
            
            avg_progress = sum(total_progress) / len(total_progress) if total_progress else 0
            user_stats.append({
                'user_id': user.id,
                'avg_progress': avg_progress
            })

        # Daily reads for chart (last 7 days)
        daily_reads = []
        date_labels = []
        for i in range(6, -1, -1):
            date = (datetime.utcnow() - timedelta(days=i)).date()
            count = ReadingProgress.query.filter(
                func.date(ReadingProgress.last_opened) == date
            ).count()
            daily_reads.append(count)
            date_labels.append(date.strftime('%Y-%m-%d'))

        return render_template('admin_dashboard.html',
                        users=total_users,
                        books=total_books,
                        total_reads=total_reads,
                        users_list=users_list,
                        users_pagination=users_pagination,
                        all_users=all_users,
                        total_bookmarks=total_bookmarks,
                        avg_books_read=avg_books_read,
                        top_books=top_books,
                        books_per_user=books_per_user,
                        currently_reading_stats=currently_reading_stats,
                        completed_books_stats=completed_books_stats,
                        active_readers=active_readers,
                        user_stats=user_stats,
                        daily_reads=daily_reads,
                        date_labels=date_labels,
                        now=datetime.utcnow(),
                        page=page,
                        per_page=per_page)

    @app.route('/admin/export_reading_by_multiple')
    @login_required
    @admin_required
    def export_reading_by_multiple():
        """Export books currently being read by multiple users to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "کتاب‌های در حال مطالعه توسط چند کاربر"
        ws.sheet_view.rightToLeft = True
        
        # Add headers
        headers = ["عنوان کتاب", "نویسنده", "تعداد خوانندگان", "تعداد کل خوانده‌شده", "لیست خوانندگان"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
        
        # Get data
        books_data = db.session.query(
            Book.title,
            Book.author,
            func.count(ReadingProgress.user_id).label('readers_count'),
            Book.read_count,
            func.group_concat(User.username.distinct().op('ORDER BY')(User.username)).label('readers')
        ).join(ReadingProgress, Book.id == ReadingProgress.book_id
        ).join(User, ReadingProgress.user_id == User.id
        ).group_by(Book.id
        ).having(func.count(ReadingProgress.user_id) > 1
        ).order_by(func.count(ReadingProgress.user_id).desc()).all()
        
        # Add data rows
        for row_num, book in enumerate(books_data, 2):
            ws.cell(row=row_num, column=1, value=book.title)
            ws.cell(row=row_num, column=2, value=book.author or 'ناشناس')
            ws.cell(row=row_num, column=3, value=book.readers_count)
            ws.cell(row=row_num, column=4, value=book.read_count)
            ws.cell(row=row_num, column=5, value=book.readers.replace(',', '، '))
        
        # Adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save and return
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name='کتابهای_در_حال_مطالعه_توسط_چند_کاربر.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    @app.route('/admin/export_completed_by_multiple')
    @login_required
    @admin_required
    def export_completed_by_multiple():
        """Export books completed by multiple users to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "کتاب‌های خوانده شده توسط چند کاربر"
        ws.sheet_view.rightToLeft = True
        
        # Add headers
        headers = ["عنوان کتاب", "نویسنده", "تعداد تکمیل‌کنندگان", "لیست خوانندگان"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
        
        # Get data (books marked completed=True in ReadingProgress).
        # نکته: مدل Book فیلد total_pages نداره (فقط ReadingProgress داره)،
        # پس مقایسه‌ی «موقعیت >= کل صفحات» اصلاً قابل اجرا نبود و AttributeError می‌داد.
        # به‌جاش از فیلد completed که دقیقاً برای همین منظور طراحی شده استفاده می‌کنیم.
        books_data = db.session.query(
            Book.title,
            Book.author,
            func.count(ReadingProgress.user_id).label('completers_count'),
            func.group_concat(User.username.distinct().op('ORDER BY')(User.username)).label('completers')
        ).join(ReadingProgress, Book.id == ReadingProgress.book_id
        ).join(User, ReadingProgress.user_id == User.id
        ).filter(ReadingProgress.completed == True
        ).group_by(Book.id
        ).having(func.count(ReadingProgress.user_id) > 1
        ).order_by(func.count(ReadingProgress.user_id).desc()).all()
        
        # Add data rows
        for row_num, book in enumerate(books_data, 2):
            ws.cell(row=row_num, column=1, value=book.title)
            ws.cell(row=row_num, column=2, value=book.author or 'ناشناس')
            ws.cell(row=row_num, column=3, value=book.completers_count)
            ws.cell(row=row_num, column=4, value=book.completers.replace(',', '، '))
        
        # Adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save and return
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name='کتابهای_تکمیل_شده_توسط_چند_کاربر.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    @app.route('/admin/export_active_readers')
    @login_required
    @admin_required
    def export_active_readers():
        """Export most active readers to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "فعال‌ترین خوانندگان"
        ws.sheet_view.rightToLeft = True
        
        # Add headers
        headers = ["نام کاربری", "تعداد کتاب‌های در حال مطالعه", "لیست کتاب‌ها"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
        
        # Get data
        active_readers = db.session.query(
            User.username,
            func.count(ReadingProgress.book_id.distinct()).label('books_reading'),
            func.group_concat(Book.title.distinct().op('ORDER BY')(Book.title)).label('book_titles')
        ).join(ReadingProgress, User.id == ReadingProgress.user_id
        ).join(Book, ReadingProgress.book_id == Book.id
        ).group_by(User.id
        ).order_by(func.count(ReadingProgress.book_id.distinct()).desc()
        ).all()
        
        # Add data rows
        for row_num, reader in enumerate(active_readers, 2):
            ws.cell(row=row_num, column=1, value=reader.username)
            ws.cell(row=row_num, column=2, value=reader.books_reading)
            ws.cell(row=row_num, column=3, value=reader.book_titles.replace(',', '، '))
        
        # Adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save and return
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name='فعال_ترین_خوانندگان.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    @app.route('/admin/reports')
    @login_required
    @admin_required
    def admin_reports():
        total_users = User.query.count()
        total_books = Book.query.count()
        total_bookmarks = Bookmark.query.count()
        users_with_progress = db.session.query(
            User.id,
            func.count(ReadingProgress.id).label('books_read')
        ).outerjoin(ReadingProgress, User.id == ReadingProgress.user_id)\
         .group_by(User.id).subquery()
         
        avg_books_read = float(db.session.query(func.avg(users_with_progress.c.books_read)).scalar() or 0)
        
        top_books = db.session.query(
            Book.title,
            func.count(Bookmark.id).label('bookmark_count')
        ).join(Bookmark)\
         .group_by(Book.id)\
         .order_by(func.desc('bookmark_count'))\
         .limit(10).all()
         
        user_stats = db.session.query(
            User.username,
            func.count(Bookmark.id).label('bookmark_count'),
            func.avg(
                # سینتکس قدیمی case([(cond, val), ...], else_=...) در SQLAlchemy 1.4+/2.0
                # حذف شده؛ فرم درست بدون لیست و با آرگومان‌های جداگانه‌ست.
                # نکته: func.min(a, b) در SQLite با دو یا چند آرگومان به‌صورت اسکالر
                # کوچک‌ترین مقدار بین آن‌ها رو برمی‌گردونه (نه aggregate min)، پس این
                # بخش برای محدود کردن درصد به حداکثر ۱۰۰ درست کار می‌کنه.
                case(
                    (ReadingProgress.total_pages > 0,
                     func.min(100.0, (ReadingProgress.position * 100.0) / ReadingProgress.total_pages)),
                    (ReadingProgress.total_pages == 0, 0.0),
                    else_=0.0
                )
            ).label('avg_progress'),
            User.last_login
        ).outerjoin(Bookmark, User.id == Bookmark.user_id)\
         .outerjoin(ReadingProgress, User.id == ReadingProgress.user_id)\
         .group_by(User.id).all()
         
        books_per_user = db.session.query(
            User.username,
            func.count(ReadingProgress.id).label('books_read')
        ).outerjoin(ReadingProgress, User.id == ReadingProgress.user_id)\
         .group_by(User.id)\
         .order_by(func.desc('books_read'))\
         .limit(10).all()
         
        return render_template('admin_reports.html',
                           total_users=total_users,
                           total_books=total_books,
                           total_bookmarks=total_bookmarks,
                           avg_books_read=round(avg_books_read, 1),
                           top_books=top_books,
                           user_stats=user_stats,
                           books_per_user=books_per_user)

    @app.route('/admin/advanced_reports')
    @login_required
    @admin_required
    def advanced_reports():
        # دسته‌بندی‌های کتاب‌ها
        category_stats = db.session.query(
            Book.category.label('name'),
            func.count(Book.id).label('book_count'),
            func.sum(Book.read_count).label('total_reads')
        ).group_by(Book.category).all()

        # نام نمایشی دسته‌بندی‌ها
        category_display_names = {
            'general': 'عمومی',
            'novel': 'رمان', 
            'science': 'علمی',
            'history': 'تاریخی',
            'religion': 'مذهبی',
            'historical_religious': 'تاریخی مذهبی'
        }

        # شمارش کتاب‌های PDF با دسته‌بندی رمان
        novel_pdf_count = db.session.query(Book).filter(
            Book.category == 'novel',
            Book.pdf_filename.isnot(None)
        ).count()
        
        # آماده‌سازی آمار دسته‌بندی‌ها برای تمپلیت
        categories_stats = []
        for cat in category_stats:
            categories_stats.append({
                'name': cat.name,
                'display_name': category_display_names.get(cat.name, cat.name.capitalize()),
                'book_count': cat.book_count,
                'total_reads': cat.total_reads or 0
            })

        # کتاب‌های برتر بر اساس تعداد نشانک‌ها
        top_books = db.session.query(
            Book.title,
            Book.author,
            Book.category,
            func.count(Bookmark.id).label('bookmark_count')
        ).outerjoin(Bookmark, Book.id == Bookmark.book_id)\
        .group_by(Book.id)\
        .order_by(func.count(Bookmark.id).desc())\
        .limit(10).all()

        # کاربران فعال
        active_users = db.session.query(
            User.username,
            func.count(distinct(ReadingProgress.book_id)).label('books_read'),
            func.sum(ReadingProgress.position).label('pages_read')
        ).join(ReadingProgress, User.id == ReadingProgress.user_id)\
        .group_by(User.id)\
        .order_by(func.count(ReadingProgress.book_id).desc())\
        .limit(10).all()

        # کتاب‌های پرمخاطب
        popular_books = db.session.query(
            Book.title,
            Book.author,
            Book.category,
            Book.created_at,  # Add this line
            func.count(ReadingProgress.id).label('read_count')
        ).join(ReadingProgress, Book.id == ReadingProgress.book_id)\
        .group_by(Book.id)\
        .order_by(func.count(ReadingProgress.id).desc())\
        .limit(10).all()

        # تعداد کل کتاب‌ها
        total_books = Book.query.count()
        
        # آمار علاقه‌مندی‌های کاربران و کتاب‌های خوانده شده
        user_interests = db.session.query(
            User.id,
            User.username,
            UserProfile.interests,
            func.group_concat(Book.title.distinct()).label('read_books_titles'),
            func.group_concat(Book.author.distinct()).label('read_books_authors'),
            func.group_concat(Book.category.distinct()).label('read_books_categories'),
            # اطلاعات کامل کتاب‌های خوانده شده
            func.group_concat(
                Book.title + '|' + Book.author + '|' + Book.category, ';'
            ).label('read_books_details')
        ).join(UserProfile, User.id == UserProfile.user_id)\
         .outerjoin(ReadingProgress, User.id == ReadingProgress.user_id)\
         .outerjoin(Book, ReadingProgress.book_id == Book.id)\
         .group_by(User.id, UserProfile.interests)
        
        # محاسبه درصد تحقق علاقه‌مندی‌ها برای هر کاربر
        user_interest_stats = []
        for user in user_interests:
            if not user.interests:
                continue
                
            interests = [i.strip() for i in user.interests.split(',') if i.strip()]
            
            # دریافت دسته‌بندی کتاب‌های خوانده شده کاربر
            user_id = db.session.query(User.id).filter(User.username == user.username).scalar()
            read_categories = db.session.query(Book.category)\
                .join(ReadingProgress, Book.id == ReadingProgress.book_id)\
                .filter(ReadingProgress.user_id == user_id)\
                .all()
            read_categories = [cat[0] for cat in read_categories if cat[0]]  # فیلتر کردن مقادیر خالی
            
            # بررسی هر علاقه‌مندی در دسته‌بندی‌های خوانده شده
            fulfilled_interests = 0
            for interest in interests:
                # تبدیل علاقه به فرمت دسته‌بندی (هر علاقه ممکن است با چند دسته مطابقت داشته باشد،
                # چون دستهٔ ترکیبی «تاریخی مذهبی» هم به علاقهٔ «تاریخی» و هم «مذهبی» مربوط می‌شود)
                interest_mapping = {
                    'تاریخی': ['history', 'historical_religious'],
                    'رمان': ['novel'],
                    'علمی': ['science'],
                    'مذهبی': ['religion', 'historical_religious'],
                    'عمومی': ['general']
                }
                categories = interest_mapping.get(interest, [interest.lower()])
                
                if any(category in read_categories for category in categories):
                    fulfilled_interests += 1
            
            fulfillment_percentage = (fulfilled_interests / len(interests)) * 100 if interests else 0
            
            # پردازش اطلاعات کتاب‌های خوانده شده
            read_books = []
            if user.read_books_details:
                book_entries = user.read_books_details.split(';')
                for entry in book_entries:
                    if '|' in entry:
                        title, author, category = entry.split('|', 2)
                        read_books.append({
                            'title': title,
                            'author': author,
                            'category': category_display_names.get(category, category.capitalize())
                        })
            
            user_interest_stats.append({
                'user_id': user.id,
                'username': user.username,
                'total_interests': len(interests),
                'fulfilled_interests': fulfilled_interests,
                'fulfillment_percentage': (fulfilled_interests / len(interests) * 100) if interests else 0,
                'interests': interests[:5],  # نمایش حداکثر 5 علاقه‌مندی
                'read_books_count': len(read_books),
                'read_books': read_books  # لیست کامل کتاب‌های خوانده شده با جزئیات
            })
        
        # مرتب‌سازی بر اساس درصد تحقق (نزولی)
        user_interest_stats.sort(key=lambda x: x['fulfillment_percentage'], reverse=True)

        return render_template('advanced_reports.html',
                        categories_stats=categories_stats,
                        top_books=top_books,
                        active_users=active_users,
                        popular_books=popular_books,
                        user_interest_stats=user_interest_stats,
                        total_books=total_books,
                        category_display_names=category_display_names,
                        novel_pdf_count=novel_pdf_count,
                        now=datetime.utcnow())  # Add current timestamp
    

    @app.route('/admin_refresh_books', methods=['POST'])
    @login_required
    def admin_refresh_books():
        if not current_user.is_admin:
            flash("دسترسی غیرمجاز", "danger")
            return redirect(url_for('user_dashboard'))
            
        stats = scan_books_folder(app.config['BASE_DIR'])
        db.session.add(ActivityLog(user_id=current_user.id, action=f"refresh_books: {stats}"))
        db.session.commit()
        
        msg_parts = []
        if stats['added_pdf']:
            msg_parts.append(f"{stats['added_pdf']} فایل PDF افزوده شد")
        if stats['removed_pdf']:
            msg_parts.append(f"{stats['removed_pdf']} فایل PDF حذف شد")
        if stats['added_txt']:
            msg_parts.append(f"{stats['added_txt']} فایل متنی افزوده شد")
        if stats['removed_txt']:
            msg_parts.append(f"{stats['removed_txt']} فایل متنی حذف شد")
        if stats['updated_txt']:
            msg_parts.append(f"{stats['updated_txt']} فایل متنی به‌روزرسانی شد")
        if stats['added_cover']:
            msg_parts.append(f"{stats['added_cover']} فایل کاور افزوده شد")
            
        flash('؛ '.join(msg_parts) + ' .' if msg_parts else "هیچ تغییر جدیدی وجود نداشت.",
              "success" if msg_parts else "info")
        return redirect(url_for('user_dashboard'))



    # ---------- Export Excel ----------
    @app.route('/admin/export_excel')
    @login_required
    def admin_export_excel():
        if not current_user.is_admin:
            flash('فقط ادمین دسترسی دارد.', 'danger')
            return redirect(url_for('user_dashboard'))

        # ---------- کمک‌تابع‌های فارسی‌سازی ----------
        PERSIAN_DIGITS = str.maketrans('0123456789', '۰۱۲۳۴۵۶۷۸۹')

        def fa_digits(s):
            return s.translate(PERSIAN_DIGITS) if s else s

        def jalali_str(dt, default=''):
            """تبدیل datetime میلادی به رشته‌ی تاریخ/ساعت شمسی با ارقام فارسی."""
            s = format_jalali(dt, default)
            return fa_digits(s) if s else default

        CATEGORY_LABELS = {
            'general': 'عمومی',
            'novel': 'رمان',
            'science': 'علمی',
            'history': 'تاریخی',
            'religion': 'مذهبی',
            'historical_religious': 'تاریخی مذهبی',
            'pdf': 'PDF',
            'txt': 'متنی',
            'epub': 'EPUB',
        }

        def category_label(cat):
            return CATEGORY_LABELS.get(cat, cat or 'نامشخص')

        users_rows, books_rows, progress_rows = [], [], []

        # Process users data
        for u in User.query.order_by(User.id).all():
            login_logs = ActivityLog.query.filter_by(user_id=u.id, action='login').order_by(ActivityLog.timestamp).all()
            logout_logs = ActivityLog.query.filter_by(user_id=u.id, action='logout').order_by(ActivityLog.timestamp).all()
            login_times = [jalali_str(l.timestamp) for l in login_logs]
            logout_times = [jalali_str(l.timestamp) for l in logout_logs]
            last_login = jalali_str(u.last_login)
            progresses = ReadingProgress.query.filter_by(user_id=u.id).all()
            books_read = [p.book.title for p in progresses if p.book]
            users_rows.append({
                'نام کاربری': u.username,
                'زمان‌های ورود': '؛ '.join(login_times),
                'زمان‌های خروج': '؛ '.join(logout_times),
                'آخرین ورود': last_login,
                'تعداد کتاب‌های خوانده‌شده': len(books_read),
                'کتاب‌های خوانده‌شده': '، '.join(books_read)
            })

        df_users = pd.DataFrame(users_rows, columns=[
            'نام کاربری', 'زمان‌های ورود', 'زمان‌های خروج',
            'آخرین ورود', 'تعداد کتاب‌های خوانده‌شده', 'کتاب‌های خوانده‌شده'
        ])

        # Process books data
        for b in Book.query.order_by(Book.id).all():
            readers = [User.query.get(p.user_id).username for p in b.progresses if User.query.get(p.user_id)]
            books_rows.append({
                'عنوان کتاب': b.title,
                'نویسنده': b.author or '',
                'دسته‌بندی': category_label(b.category),
                'تعداد خوانندگان': len(readers),
                'خوانندگان': '، '.join(readers)
            })

        df_books = pd.DataFrame(books_rows, columns=[
            'عنوان کتاب', 'نویسنده', 'دسته‌بندی', 'تعداد خوانندگان', 'خوانندگان'
        ])

        # Process reading progress
        for p in ReadingProgress.query.order_by(ReadingProgress.id).all():
            progress_rows.append({
                'شناسه کاربر': p.user_id,
                'نام کاربری': User.query.get(p.user_id).username if User.query.get(p.user_id) else '',
                'شناسه کتاب': p.book_id,
                'عنوان کتاب': Book.query.get(p.book_id).title if Book.query.get(p.book_id) else '',
                'موقعیت مطالعه': p.position,
                'آخرین بازدید': jalali_str(p.last_opened)
            })

        df_progress = pd.DataFrame(progress_rows, columns=[
            'شناسه کاربر', 'نام کاربری', 'شناسه کتاب', 'عنوان کتاب', 'موقعیت مطالعه', 'آخرین بازدید'
        ])

        # ---------- Sales report data ----------
        paid_purchases = Purchase.query.filter_by(status='paid').order_by(Purchase.paid_at).all()

        # میانگین فروش هفتگی و ماهانه
        weekly_rows, monthly_rows = [], []
        if paid_purchases:
            df_sales = pd.DataFrame([{
                'paid_at': p.paid_at,
                'amount': p.amount
            } for p in paid_purchases if p.paid_at])

            if not df_sales.empty:
                df_sales['week_period'] = df_sales['paid_at'].dt.to_period('W')
                weekly_group = df_sales.groupby('week_period').agg(
                    count=('amount', 'count'), revenue=('amount', 'sum')
                )
                weekly_rows.append({
                    'شاخص': 'میانگین تعداد فروش هفتگی (تعداد خرید در هفته)',
                    'مقدار': round(weekly_group['count'].mean(), 2)
                })
                weekly_rows.append({
                    'شاخص': 'میانگین مبلغ فروش هفتگی (تومان)',
                    'مقدار': round(weekly_group['revenue'].mean(), 2)
                })

                df_sales['month_period'] = df_sales['paid_at'].dt.to_period('M')
                monthly_group = df_sales.groupby('month_period').agg(
                    count=('amount', 'count'), revenue=('amount', 'sum')
                )
                monthly_rows.append({
                    'شاخص': 'میانگین تعداد فروش ماهانه (تعداد خرید در ماه)',
                    'مقدار': round(monthly_group['count'].mean(), 2)
                })
                monthly_rows.append({
                    'شاخص': 'میانگین مبلغ فروش ماهانه (تومان)',
                    'مقدار': round(monthly_group['revenue'].mean(), 2)
                })

        sales_summary_rows = weekly_rows + monthly_rows
        sales_summary_rows.append({
            'شاخص': 'تعداد کل کتاب‌های موجود در کتابخانه',
            'مقدار': Book.query.count()
        })
        sales_summary_rows.append({
            'شاخص': 'تعداد کتاب‌های منحصربه‌فردی که حداقل یک‌بار خریداری شده‌اند',
            'مقدار': db.session.query(func.count(func.distinct(Purchase.book_id)))
                .filter(Purchase.status == 'paid').scalar() or 0
        })
        sales_summary_rows.append({
            'شاخص': 'تعداد کل خریدهای موفق (رکورد)',
            'مقدار': len(paid_purchases)
        })
        sales_summary_rows.append({
            'شاخص': 'مجموع درآمد کل (تومان)',
            'مقدار': sum(p.amount for p in paid_purchases)
        })
        sales_summary_rows.append({
            'شاخص': 'میانگین دفعات خرید هر کتاب (از میان کتاب‌های موجود)',
            'مقدار': round(len(paid_purchases) / Book.query.count(), 2) if Book.query.count() else 0
        })
        sales_summary_rows.append({
            'شاخص': 'تاریخ و ساعت تهیه گزارش',
            'مقدار': jalali_str(datetime.utcnow())
        })
        df_sales_summary = pd.DataFrame(sales_summary_rows, columns=['شاخص', 'مقدار'])

        # تعداد دفعات خرید هر کتاب نسبت به موجودی کتابخانه
        book_purchase_counts = dict(
            db.session.query(Purchase.book_id, func.count(Purchase.id))
            .filter(Purchase.status == 'paid')
            .group_by(Purchase.book_id).all()
        )
        book_revenue = dict(
            db.session.query(Purchase.book_id, func.sum(Purchase.amount))
            .filter(Purchase.status == 'paid')
            .group_by(Purchase.book_id).all()
        )
        book_inventory_rows = []
        for b in Book.query.order_by(Book.id).all():
            book_inventory_rows.append({
                'شناسه کتاب': b.id,
                'عنوان کتاب': b.title,
                'نویسنده': b.author or '',
                'دسته‌بندی': category_label(b.category),
                'قیمت (تومان)': b.price,
                'تعداد خرید': book_purchase_counts.get(b.id, 0),
                'درآمد کل (تومان)': book_revenue.get(b.id, 0) or 0
            })
        df_book_inventory = pd.DataFrame(book_inventory_rows, columns=[
            'شناسه کتاب', 'عنوان کتاب', 'نویسنده', 'دسته‌بندی', 'قیمت (تومان)', 'تعداد خرید', 'درآمد کل (تومان)'
        ]).sort_values('تعداد خرید', ascending=False)

        # محبوبیت بر اساس نام کتاب (تکراری از df_book_inventory اما فقط خریداری‌شده‌ها)
        df_popular_books = df_book_inventory[df_book_inventory['تعداد خرید'] > 0][
            ['عنوان کتاب', 'نویسنده', 'دسته‌بندی', 'تعداد خرید', 'درآمد کل (تومان)']
        ].reset_index(drop=True)

        # محبوبیت بر اساس نویسنده
        author_rows = []
        for author, group in df_book_inventory.groupby('نویسنده'):
            if not author:
                continue
            author_rows.append({
                'نویسنده': author,
                'تعداد خرید': int(group['تعداد خرید'].sum()),
                'درآمد کل (تومان)': int(group['درآمد کل (تومان)'].sum()),
                'تعداد کتاب‌های متفاوت خریداری‌شده': int((group['تعداد خرید'] > 0).sum())
            })
        df_popular_authors = pd.DataFrame(author_rows, columns=[
            'نویسنده', 'تعداد خرید', 'درآمد کل (تومان)', 'تعداد کتاب‌های متفاوت خریداری‌شده'
        ]).sort_values('تعداد خرید', ascending=False)

        # محبوبیت بر اساس دسته‌بندی
        category_rows = []
        for category, group in df_book_inventory.groupby('دسته‌بندی'):
            category_rows.append({
                'دسته‌بندی': category,
                'تعداد خرید': int(group['تعداد خرید'].sum()),
                'درآمد کل (تومان)': int(group['درآمد کل (تومان)'].sum()),
                'تعداد کتاب‌های متفاوت خریداری‌شده': int((group['تعداد خرید'] > 0).sum())
            })
        df_popular_categories = pd.DataFrame(category_rows, columns=[
            'دسته‌بندی', 'تعداد خرید', 'درآمد کل (تومان)', 'تعداد کتاب‌های متفاوت خریداری‌شده'
        ]).sort_values('تعداد خرید', ascending=False)

        # تعداد کتاب خریداری‌شده توسط هر کاربر
        user_purchase_counts = dict(
            db.session.query(Purchase.user_id, func.count(Purchase.id))
            .filter(Purchase.status == 'paid')
            .group_by(Purchase.user_id).all()
        )
        user_spend = dict(
            db.session.query(Purchase.user_id, func.sum(Purchase.amount))
            .filter(Purchase.status == 'paid')
            .group_by(Purchase.user_id).all()
        )
        user_purchase_rows = []
        for u in User.query.order_by(User.id).all():
            books_bought = user_purchase_counts.get(u.id, 0)
            if books_bought == 0:
                continue
            user_purchase_rows.append({
                'شناسه کاربر': u.id,
                'نام کاربری': u.username,
                'تعداد کتاب خریداری‌شده': books_bought,
                'مجموع مبلغ خرید (تومان)': user_spend.get(u.id, 0) or 0
            })
        df_user_purchases = pd.DataFrame(user_purchase_rows, columns=[
            'شناسه کاربر', 'نام کاربری', 'تعداد کتاب خریداری‌شده', 'مجموع مبلغ خرید (تومان)'
        ]).sort_values('تعداد کتاب خریداری‌شده', ascending=False)

        # Create Excel file
        SHEETS = [
            ('کاربران', df_users),
            ('کتاب‌ها', df_books),
            ('پیشرفت مطالعه', df_progress),
            ('خلاصه فروش', df_sales_summary),
            ('کتاب‌ها و تعداد خرید', df_book_inventory),
            ('محبوب‌ترین کتاب‌ها', df_popular_books),
            ('محبوب‌ترین نویسندگان', df_popular_authors),
            ('محبوب‌ترین دسته‌بندی‌ها', df_popular_categories),
            ('خرید کاربران', df_user_purchases),
        ]

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, df in SHEETS:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            header_font = Font(bold=True)
            for sheet_name, df in SHEETS:
                ws = writer.sheets[sheet_name]
                ws.sheet_view.rightToLeft = True
                for cell in ws[1]:
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                # عرض ستون‌ها را متناسب با محتوا تنظیم می‌کنیم
                for col_idx, col_name in enumerate(df.columns, start=1):
                    max_len = max(
                        [len(str(col_name))] + [len(str(v)) for v in df.iloc[:, col_idx - 1].astype(str)]
                    ) if len(df) else len(str(col_name))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name='گزارش_فروش_کتابخانه.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    @app.route('/download_report')
    @login_required
    @admin_required
    def download_report():
        return admin_export_excel()

    # نکته: یک تابع دیگر به نام admin_reports_redirect قبلاً روی همین URL
    # ('/admin/reports') ثبت شده بود. چون هر دو رول روی یک مسیر یکسان بودن،
    # Werkzeug فقط اولین رول ثبت‌شده (admin_reports) رو مچ می‌کرد و این تابع
    # دوم عملاً هیچ‌وقت اجرا نمی‌شد (کد مرده). چون کار متفاوتی نسبت به
    # admin_reports انجام نمی‌داد، حذف شد. اگر نیاز به یک مسیر قدیمی/alias
    # جدا برای سازگاری با نسخه‌های قبلی هست، باید URL متفاوتی (مثلاً
    # '/reports') بهش داده بشه.

    @app.route('/profile', methods=['GET', 'POST'])
    @login_required
    def profile():
        form = ProfileForm()
        profile = UserProfile.query.filter_by(user_id=current_user.id).first()
        
        if not profile:
            profile = UserProfile(user_id=current_user.id)
            db.session.add(profile)
            db.session.commit()
            
        if form.validate_on_submit():
            try:
                if form.profile_image.data:
                    file = form.profile_image.data
                    if file.filename != '':
                        if profile.profile_image:
                            try:
                                old_image_path = os.path.join(app.root_path, 'static', profile.profile_image)
                                if os.path.exists(old_image_path):
                                    os.remove(old_image_path)
                            except Exception as e:
                                app.logger.error(f'Error removing old profile image: {str(e)}')
                        
                        try:
                            filename = unicodedata.normalize('NFC', file.filename)  # نرمال‌سازی نام فایل
                            safe_filename = secure_filename(filename)  # حذف کاراکترهای نامناسب
                            if not safe_filename:
                                safe_filename = f"profile_{uuid.uuid4().hex}"
                            unique_filename = f"{uuid.uuid4().hex}_{safe_filename}"
                            
                            upload_folder = os.path.join(app.root_path, 'static', 'profile_images')
                            os.makedirs(upload_folder, exist_ok=True)
                            file_path = os.path.join(upload_folder, unique_filename)
                            file.save(file_path)
                            profile.profile_image = os.path.join('profile_images', unique_filename).replace('\\', '/')
                        except Exception as e:
                            app.logger.error(f'Error saving profile image: {str(e)}')
                            flash('خطا در ذخیره تصویر پروفایل', 'danger')
                
                try:
                    profile.full_name = form.full_name.data
                    profile.location = form.location.data
                    profile.job = form.job.data
                    profile.education = form.education.data
                    profile.field_of_study = form.field_of_study.data
                    profile.bio = form.bio.data
                    profile.phone_number = form.phone_number.data
                    profile.social_media = form.social_media.data
                    profile.interests = form.interests.data

                    # ثبت/تغییر ایمیل از صفحه پروفایل — مخصوصاً برای کاربرانی
                    # که از قبل حساب داشتند و هنوز ایمیلی ثبت نکرده بودند.
                    new_email = (form.email.data or '').strip().lower()
                    if new_email and new_email != (current_user.email or ''):
                        if not EMAIL_RE.match(new_email):
                            flash('ایمیل وارد شده معتبر نیست.', 'danger')
                            return render_template('profile.html', form=form, profile=profile)
                        existing = User.query.filter(User.email == new_email, User.id != current_user.id).first()
                        if existing:
                            flash('این ایمیل قبلاً برای حساب دیگری ثبت شده.', 'danger')
                            return render_template('profile.html', form=form, profile=profile)
                        current_user.email = new_email
                        current_user.email_verified = False
                        send_verification_email(app, current_user)
                        flash('ایمیل به‌روزرسانی شد. یک ایمیل تأیید جدید برای شما ارسال شد.', 'info')

                    db.session.commit()
                    flash('پروفایل با موفقیت به‌روزرسانی شد', 'success')
                    return redirect(url_for('profile'))
                except Exception as e:
                    db.session.rollback()
                    app.logger.error(f'Error updating profile: {str(e)}')
                    flash('خطا در به‌روزرسانی پروفایل', 'danger')
                    
            except Exception as e:
                db.session.rollback()
                app.logger.error(f'Unexpected error in profile update: {str(e)}')
                flash('خطای غیرمنتظره در به‌روزرسانی پروفایل', 'danger')
                
        # Pre-fill form with existing data
        form.full_name.data = profile.full_name
        form.location.data = profile.location
        form.job.data = profile.job
        form.education.data = profile.education
        form.field_of_study.data = profile.field_of_study
        form.bio.data = profile.bio
        form.phone_number.data = profile.phone_number
        form.social_media.data = profile.social_media
        form.interests.data = profile.interests
        if request.method == 'GET':
            form.email.data = current_user.email
        
        # Get unread messages count
        unread_count = Message.query.filter_by(
            recipient_id=current_user.id,
            is_read=False
        ).count()
        
        # Get admin user for messaging
        admin_user = User.query.filter_by(is_admin=True).first()
        
        return render_template('profile.html', 
                             form=form, 
                             profile=profile, 
                             unread_count=unread_count,
                             admin_user=admin_user)
    
    @app.route('/send_message/<int:recipient_id>', methods=['GET', 'POST'])
    @login_required
    def send_message(recipient_id):
        recipient = User.query.get_or_404(recipient_id)
        form = MessageForm()
        
        if form.validate_on_submit():
            message = Message(
                sender_id=current_user.id,
                recipient_id=recipient.id,
                content=form.content.data
            )
            db.session.add(message)
            db.session.commit()
            flash('پیام شما ارسال شد', 'success')
            return redirect(url_for('view_messages', user_id=recipient.id))
            
        return render_template('send_message.html', 
                             title='ارسال پیام',
                             form=form, 
                             recipient=recipient)
    
    @app.route('/messages/<int:user_id>')
    @login_required
    def view_messages(user_id):
        other_user = User.query.get_or_404(user_id)
        
        # Mark messages as read when viewing
        Message.query.filter_by(
            sender_id=other_user.id,
            recipient_id=current_user.id,
            is_read=False
        ).update({'is_read': True})
        db.session.commit()
        
        # Get all messages between current user and other user
        messages = Message.query.filter(
            ((Message.sender_id == current_user.id) & (Message.recipient_id == other_user.id)) |
            ((Message.sender_id == other_user.id) & (Message.recipient_id == current_user.id))
        ).order_by(Message.timestamp.asc()).all()
        
        return render_template('messages.html', 
                             messages=messages, 
                             other_user=other_user,
                             form=MessageForm())
    
    @app.route('/messages')
    @login_required
    def messages():
        # Get all conversations
        sent_messages = Message.query.filter_by(sender_id=current_user.id).all()
        received_messages = Message.query.filter_by(recipient_id=current_user.id).all()
        
        # Get unique user IDs from conversations
        user_ids = set()
        for msg in sent_messages:
            user_ids.add(msg.recipient_id)
        for msg in received_messages:
            user_ids.add(msg.sender_id)
        
        # Get user objects and last message for each conversation
        conversations = []
        for uid in user_ids:
            user = User.query.get(uid)
            if user:
                last_message = Message.query.filter(
                    ((Message.sender_id == current_user.id) & (Message.recipient_id == uid)) |
                    ((Message.sender_id == uid) & (Message.recipient_id == current_user.id))
                ).order_by(Message.timestamp.desc()).first()
                
                unread_count = Message.query.filter_by(
                    sender_id=uid,
                    recipient_id=current_user.id,
                    is_read=False
                ).count()
                
                conversations.append({
                    'user': user,
                    'last_message': last_message,
                    'unread_count': unread_count
                })
        
        # Sort conversations by last message timestamp
        conversations.sort(key=lambda x: x['last_message'].timestamp if x['last_message'] else datetime.min, reverse=True)
        
        # Get admin user to show in the template
        admin_user = User.query.filter_by(is_admin=True).first()
        
        return render_template('conversations.html', 
                            conversations=conversations,
                            admin_user=admin_user)
    
    @app.route('/message/<int:message_id>/delete', methods=['POST'])
    @login_required
    def delete_message(message_id):
        message = Message.query.get_or_404(message_id)
        if message.sender_id != current_user.id and message.recipient_id != current_user.id:
            abort(403)
        
        db.session.delete(message)
        db.session.commit()
        flash('پیام حذف شد', 'success')
        return redirect(url_for('messages'))

    # تعداد تلاش‌های زیاد روی فرم‌های ورود/ثبت‌نام/فراموشی رمز (ناشی از
    # limiter.limit روی این روت‌ها) به‌جای صفحه‌ی خطای خام، پیام فارسی و
    # بازگشت به همون صفحه رو نشون می‌ده.
    @app.errorhandler(429)
    def ratelimit_handler(e):
        flash('تعداد تلاش‌های شما بیش از حد مجاز بوده. لطفاً چند دقیقه دیگر دوباره امتحان کنید.', 'danger')
        return redirect(request.referrer or url_for('login')), 429

    # ---------- صفحات خطای اختصاصی ----------
    # جلوگیری از نمایش صفحه‌ی خام Werkzeug/Flask به کاربر (که هم زشته و هم
    # می‌تونه جزئیات فنی سرور رو لو بده).
    @app.errorhandler(404)
    def not_found_error(e):
        return render_template('errors/404.html'), 404

    @app.errorhandler(403)
    def forbidden_error(e):
        return render_template('errors/403.html'), 403

    @app.errorhandler(413)
    def payload_too_large_error(e):
        return render_template('errors/413.html'), 413

    @app.errorhandler(500)
    def internal_error(e):
        # هر خطای پیش‌بینی‌نشده باید رول‌بک بشه تا session دیتابیس در حالت
        # خراب برای request بعدی نمونه.
        db.session.rollback()
        app.logger.error(f"Internal server error: {e}", exc_info=True)
        return render_template('errors/500.html'), 500

    # ---------- هدرهای امنیتی HTTP ----------
    # محافظت پایه در برابر clickjacking (X-Frame-Options)، MIME sniffing
    # (X-Content-Type-Options)، نشت referrer، و XSS/تزریق (Content-Security-Policy).
    # این‌ها به‌صورت دستی اضافه شدن تا وابستگی جدیدی (مثل flask-talisman) به
    # requirements.txt اضافه نشه.
    @app.after_request
    def set_security_headers(response):
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['X-Frame-Options'] = 'DENY'
        response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
        response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
        # CSP: اسکریپت/استایل فقط از همین دامنه (unsafe-inline برای سازگاری با
        # استایل/اسکریپت‌های inline موجود در قالب‌ها لازمه؛ در آینده با انتقال
        # این‌ها به فایل‌های جدا می‌شه سخت‌گیرانه‌ترش کرد).
        response.headers['Content-Security-Policy'] = (
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline'; "
            "style-src 'self' 'unsafe-inline'; "
            "img-src 'self' data: blob:; "
            "font-src 'self' data:; "
            "frame-ancestors 'none'; "
            "object-src 'none'"
        )
        if app.config.get('IS_PRODUCTION') and request.is_secure:
            # HSTS فقط وقتی معنی داره که واقعاً روی HTTPS اجرا بشه.
            response.headers['Strict-Transport-Security'] = 'max-age=63072000; includeSubDomains'
        return response

    return app

if __name__ == '__main__':
    app = create_app()
    # debug=True دیباگر Werkzeug رو فعال می‌کنه که اجازه‌ی اجرای کد دلخواه
    # از طریق مرورگر رو می‌ده — هرگز نباید در پروداکشن روشن باشه.
    # پیش‌فرض False است؛ فقط با ست کردن صریح FLASK_DEBUG=1 در محیط توسعه فعال می‌شه.
    debug_mode = os.environ.get('FLASK_DEBUG', '0') == '1'
    app.run(debug=debug_mode)
